from fastapi import FastAPI, APIRouter, HTTPException, Header, File, UploadFile, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from dateutil.relativedelta import relativedelta
from enum import Enum
import re
import random
import asyncio
import aiohttp
import json
import json
import jwt
import time
from passlib.context import CryptContext
from passlib.hash import bcrypt

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Authentication Configuration
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-here-please-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# User Roles
class UserRole(str, Enum):
    ADMIN = "admin"
    MANAGER = "manager" 
    USER = "user"  # Customer

# Create the main app without a prefix
app = FastAPI(title="FPT Bill Manager API", version="1.0.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Enums
class Gateway(str, Enum):
    FPT = "FPT"
    SHOPEE = "SHOPEE"

class ProviderRegion(str, Enum):
    MIEN_BAC = "MIEN_BAC"
    MIEN_NAM = "MIEN_NAM"
    HCMC = "HCMC"

class BillStatus(str, Enum):
    AVAILABLE = "AVAILABLE"
    PENDING = "PENDING"
    SOLD = "SOLD"
    CROSSED = "CROSSED"  # Đã gạch - customer không nợ cước
    ERROR = "ERROR"

class CustomerType(str, Enum):
    INDIVIDUAL = "INDIVIDUAL"
    AGENT = "AGENT"

class PaymentMethod(str, Enum):
    CASH = "CASH"
    BANK_TRANSFER = "BANK_TRANSFER"
    OTHER = "OTHER"

# Authentication Models
class UserCreate(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    email: EmailStr
    phone: Optional[str] = Field(None, min_length=10, max_length=15)
    password: str = Field(..., min_length=6)
    full_name: str = Field(..., min_length=1, max_length=100)
    role: UserRole = UserRole.USER

class UserLogin(BaseModel):
    login: str  # Can be username, email, or phone
    password: str

class UserResponse(BaseModel):
    id: str
    username: str
    email: str
    phone: Optional[str]
    full_name: str
    role: UserRole
    is_active: bool
    created_at: datetime
    last_login: Optional[datetime] = None

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class UserUpdate(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[EmailStr] = None

class PasswordChange(BaseModel):
    current_password: str
    new_password: str = Field(..., min_length=6)

class CardType(str, Enum):
    VISA = "VISA"
    MASTERCARD = "MASTERCARD"
    JCB = "JCB"
    AMEX = "AMEX"

class CardStatus(str, Enum):
    PAID_OFF = "Đã đáo"  # Đã đáo
    NEED_PAYMENT = "Cần đáo"  # Cần đáo  
    NOT_DUE = "Chưa đến hạn"  # Chưa đến hạn
    OVERDUE = "Quá Hạn"  # Quá hạn (red alert)

# Models
class Bill(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    gateway: Gateway
    customer_code: str
    provider_region: ProviderRegion
    provider_name: Optional[str] = None
    full_name: Optional[str] = None
    address: Optional[str] = None
    amount: Optional[float] = None
    billing_cycle: Optional[str] = None
    raw_status: Optional[str] = None
    status: BillStatus = BillStatus.AVAILABLE
    error_code: Optional[str] = None
    error_message: Optional[str] = None
    meta: Optional[Dict[str, Any]] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_checked: Optional[datetime] = None

class InventoryItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    bill_id: str
    bill: Optional[Bill] = None
    note: Optional[str] = None
    added_by: Optional[str] = "system"  # user who added to inventory
    batch_id: Optional[str] = None  # group bills added together
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Customer(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: CustomerType = CustomerType.INDIVIDUAL
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    is_active: bool = True
    # Transaction tracking
    total_transactions: int = 0
    total_value: float = 0.0
    total_bills: int = 0
    total_cards: int = 0
    total_profit_generated: float = 0.0
    # Metadata
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    type: CustomerType = CustomerType.INDIVIDUAL
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None

class CustomerUpdate(BaseModel):
    type: Optional[CustomerType] = None
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    is_active: Optional[bool] = None
    notes: Optional[str] = None

class CustomerStats(BaseModel):
    total_customers: int
    individual_customers: int
    agent_customers: int
    active_customers: int
    total_customer_value: float

class CreditCard(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str  # Link to customer who owns this card
    customer_name: str  # Customer name for easy access
    card_number: str  # Full card number (will be masked in display)
    cardholder_name: str  # Name on the card
    bank_name: str
    card_type: CardType
    expiry_date: str  # MM/YY format
    ccv: str  # 3-4 digits
    statement_date: int  # Day of month (1-31)
    payment_due_date: int  # Day of month (1-31)
    credit_limit: float  # Hạng mức
    status: CardStatus
    notes: Optional[str] = None
    # Cycle tracking fields
    current_cycle_month: Optional[str] = None  # MM/YYYY format
    last_payment_date: Optional[datetime] = None  # Ngày đáo gần nhất trong chu kỳ hiện tại
    cycle_payment_count: int = 0  # Số lần đáo trong chu kỳ hiện tại
    total_cycles: int = 0  # Tổng số chu kỳ đã trải qua
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CreditCardCreate(BaseModel):
    customer_id: str
    card_number: str
    cardholder_name: str
    bank_name: str
    card_type: CardType
    expiry_date: str  # MM/YY
    ccv: str
    statement_date: int
    payment_due_date: int
    credit_limit: float
    status: CardStatus = CardStatus.NOT_DUE
    notes: Optional[str] = None

class CreditCardUpdate(BaseModel):
    card_number: Optional[str] = None
    cardholder_name: Optional[str] = None
    bank_name: Optional[str] = None
    card_type: Optional[CardType] = None
    expiry_date: Optional[str] = None
    ccv: Optional[str] = None
    statement_date: Optional[int] = None
    payment_due_date: Optional[int] = None
    credit_limit: Optional[float] = None
    status: Optional[CardStatus] = None
    notes: Optional[str] = None

class CreditCardStats(BaseModel):
    total_cards: int
    paid_off_cards: int
    need_payment_cards: int
    not_due_cards: int
    total_credit_limit: float

class CreditCardPaymentMethod(str, Enum):
    POS = "POS"  # Thanh toán qua POS
    BILL = "BILL"  # Thanh toán qua bill điện

class CreditCardTransaction(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    card_id: str  # Link to credit card
    customer_id: str  # Link to customer
    transaction_group_id: str  # Nhóm các transactions cùng một lần đáo
    transaction_type: str = "CREDIT_CARD_PAYMENT"  # Loại giao dịch
    payment_method: CreditCardPaymentMethod  # POS hoặc BILL
    total_amount: float  # Tổng số tiền đáo
    profit_pct: float  # % lợi nhuận
    profit_value: float  # Giá trị lợi nhuận
    payback: float  # Số tiền trả khách
    bill_ids: Optional[List[str]] = []  # Danh sách bill_ids nếu dùng phương thức BILL
    notes: Optional[str] = None
    status: str = "COMPLETED"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CreditCardTransactionCreate(BaseModel):
    payment_method: CreditCardPaymentMethod
    total_amount: Optional[float] = None  # For POS method
    profit_pct: float
    bill_ids: Optional[List[str]] = []  # For BILL method
    notes: Optional[str] = None

class ActivityType(str, Enum):
    CUSTOMER_CREATE = "CUSTOMER_CREATE"      # Thêm khách hàng
    CUSTOMER_DELETE = "CUSTOMER_DELETE"      # Xóa khách hàng
    BILL_CREATE = "BILL_CREATE"              # Thêm bill
    BILL_SALE = "BILL_SALE"                  # Bán bill
    BILL_DELETE = "BILL_DELETE"              # Xóa bill
    CARD_CREATE = "CARD_CREATE"              # Thêm thẻ
    CARD_DELETE = "CARD_DELETE"              # Xóa thẻ
    CARD_PAYMENT_POS = "CARD_PAYMENT_POS"    # Đáo thẻ POS
    CARD_PAYMENT_BILL = "CARD_PAYMENT_BILL"  # Đáo thẻ BILL
    SYSTEM_ERROR = "SYSTEM_ERROR"            # Lỗi hệ thống

class Activity(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    type: ActivityType
    title: str                               # "Đáo thẻ ****1234 - 5M VND"
    description: Optional[str] = None        # Details
    customer_id: Optional[str] = None        # Link to customer
    customer_name: Optional[str] = None      # For display
    amount: Optional[float] = None           # Transaction amount
    status: str = "SUCCESS"                  # SUCCESS, ERROR, WARNING
    metadata: Optional[Dict[str, Any]] = None # Additional data
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ActivityCreate(BaseModel):
    type: ActivityType
    title: str
    description: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    amount: Optional[float] = None
    status: str = "SUCCESS"
    metadata: Optional[Dict[str, Any]] = None

class Sale(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    transaction_type: str = "ELECTRIC_BILL"  # ELECTRIC_BILL or CREDIT_CARD
    total: float
    profit_pct: float
    profit_value: float
    payback: float
    method: PaymentMethod
    status: str = "COMPLETED"
    notes: Optional[str] = None
    bill_ids: List[str]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SaleCreate(BaseModel):
    customer_id: str
    bill_ids: List[str]
    profit_pct: float
    method: PaymentMethod
    notes: Optional[str] = None

# Request/Response Models
class AddToInventoryRequest(BaseModel):
    bill_ids: List[str]
    note: Optional[str] = None
    batch_name: Optional[str] = None

class InventoryStats(BaseModel):
    total_bills: int  # Bills in inventory
    available_bills: int
    pending_bills: int
    sold_bills: int
    total_value: float
    total_bills_in_system: int  # All bills in bills collection

class InventoryResponse(BaseModel):
    id: str
    bill_id: str
    customer_code: str
    full_name: Optional[str] = None
    address: Optional[str] = None
    amount: Optional[float] = None
    billing_cycle: Optional[str] = None
    provider_region: str
    status: str
    note: Optional[str] = None
    batch_id: Optional[str] = None
    created_at: datetime

# Request/Response Models
class CheckBillRequest(BaseModel):
    gateway: Gateway
    provider_region: ProviderRegion
    codes: List[str]

class CheckBillResult(BaseModel):
    customer_code: str
    full_name: Optional[str] = None
    address: Optional[str] = None
    amount: Optional[float] = None
    billing_cycle: Optional[str] = None
    status: str
    errors: Optional[Dict[str, str]] = None
    bill_id: Optional[str] = None  # Add bill_id for inventory

class CheckBillResponse(BaseModel):
    items: List[CheckBillResult]
    summary: Dict[str, int]

class DashboardStats(BaseModel):
    total_bills: int
    available_bills: int
    sold_bills: int
    total_customers: int
    total_revenue: float
    recent_activities: List[Dict[str, Any]]

class WebhookPayload(BaseModel):
    bills: List[Dict[str, Any]]
    timestamp: int
    request_id: str
    webhook_url: str
    execution_mode: str

class BillCreate(BaseModel):
    customer_code: str
    provider_region: ProviderRegion
    full_name: Optional[str] = None
    address: Optional[str] = None
    amount: Optional[float] = None
    billing_cycle: Optional[str] = None  # Format: MM/YYYY
    status: BillStatus = BillStatus.AVAILABLE

# Authentication Utility Functions
def hash_password(password: str) -> str:
    """Hash password using bcrypt"""
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify password against hash"""
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    """Create JWT access token"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(token: str) -> Optional[dict]:
    """Verify JWT token and return payload"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        return payload
    except jwt.PyJWTError:
        return None

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    """Get current user from JWT token"""
    token = credentials.credentials
    payload = verify_token(token)
    
    if payload is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    user_id: str = payload.get("sub")
    if user_id is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid authentication credentials",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Get user from database
    user = await db.users.find_one({"id": user_id})
    if user is None:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="User not found",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    return user

def require_role(required_roles: List[UserRole]):
    """Dependency to require specific user roles"""
    def role_checker(current_user: dict = Depends(get_current_user)):
        user_role = current_user.get("role")
        if user_role not in required_roles:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Insufficient permissions"
            )
        return current_user
    return role_checker

# Specific role dependencies
admin_required = Depends(require_role([UserRole.ADMIN]))
manager_or_admin_required = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER]))
authenticated_user = Depends(get_current_user)

# Credit Card Cycle Business Logic Functions
def get_current_cycle_month(current_date: datetime = None) -> str:
    """Get current cycle month in MM/YYYY format"""
    if not current_date:
        current_date = datetime.now(timezone.utc)
    return current_date.strftime("%m/%Y")

def get_next_cycle_date(statement_date: int, current_date: datetime = None) -> datetime:
    """Get next cycle statement date"""
    if not current_date:
        current_date = datetime.now(timezone.utc)
    
    # Create next statement date
    if current_date.day < statement_date:
        # Statement date hasn't passed this month
        next_cycle = current_date.replace(day=statement_date, hour=0, minute=0, second=0, microsecond=0)
    else:
        # Statement date passed, go to next month
        next_month = current_date + relativedelta(months=1)
        next_cycle = next_month.replace(day=statement_date, hour=0, minute=0, second=0, microsecond=0)
    
    return next_cycle

def get_payment_due_date(statement_date: int, payment_due_date: int, current_date: datetime = None) -> datetime:
    """Get payment due date for current cycle"""
    if not current_date:
        current_date = datetime.now(timezone.utc)
    
    # Determine current cycle statement date
    if current_date.day >= statement_date:
        # We're in current cycle
        cycle_month = current_date
    else:
        # We're in previous cycle
        cycle_month = current_date - relativedelta(months=1)
    
    # Payment due is usually next month
    due_month = cycle_month + relativedelta(months=1)
    
    try:
        due_date = due_month.replace(day=payment_due_date, hour=23, minute=59, second=59, microsecond=0)
    except ValueError:
        # Handle month with fewer days
        due_date = due_month.replace(day=28, hour=23, minute=59, second=59, microsecond=0)
    
    return due_date

def calculate_card_status_realtime(card: dict, current_date: datetime = None) -> CardStatus:
    """Calculate real-time card status based on cycle logic"""
    if not current_date:
        current_date = datetime.now(timezone.utc)
    
    statement_date = card.get("statement_date", 1)
    payment_due_date = card.get("payment_due_date", 15)
    current_cycle_month = get_current_cycle_month(current_date)
    
    # Get key dates
    next_statement = get_next_cycle_date(statement_date, current_date)
    payment_due = get_payment_due_date(statement_date, payment_due_date, current_date)
    
    # Check if we're in current cycle
    card_cycle_month = card.get("current_cycle_month")
    last_payment_date = card.get("last_payment_date")
    
    # Determine current status
    if current_date < next_statement:
        # Before next statement date
        if card_cycle_month == current_cycle_month and last_payment_date:
            return CardStatus.PAID_OFF  # Đã đáo trong chu kỳ này
        elif current_date > payment_due + timedelta(days=7):
            return CardStatus.NOT_DUE  # Reset to next cycle after grace period
        elif current_date > payment_due:
            return CardStatus.OVERDUE  # Quá hạn (grace period)
        else:
            return CardStatus.NEED_PAYMENT  # Cần đáo
    else:
        # After statement date - new cycle begins
        if card_cycle_month == current_cycle_month and last_payment_date:
            return CardStatus.PAID_OFF  # Already paid this cycle
        else:
            return CardStatus.NEED_PAYMENT  # New cycle, needs payment

async def update_card_cycle_status(card_id: str, current_date: datetime = None):
    """Update card cycle status and tracking fields"""
    if not current_date:
        current_date = datetime.now(timezone.utc)
    
    card = await db.credit_cards.find_one({"id": card_id})
    if not card:
        return None
    
    current_cycle_month = get_current_cycle_month(current_date)
    new_status = calculate_card_status_realtime(card, current_date)
    
    # Check if we've entered a new cycle
    if card.get("current_cycle_month") != current_cycle_month:
        # New cycle detected - reset cycle fields
        update_fields = {
            "status": new_status.value,
            "current_cycle_month": current_cycle_month,
            "cycle_payment_count": 0,
            "total_cycles": card.get("total_cycles", 0) + 1,
            "updated_at": current_date
        }
        
        # Only reset last_payment_date if it's from previous cycle
        if card.get("current_cycle_month") and card.get("current_cycle_month") != current_cycle_month:
            update_fields["last_payment_date"] = None
    else:
        # Same cycle - just update status
        update_fields = {
            "status": new_status.value,
            "updated_at": current_date
        }
    
    # Update card
    await db.credit_cards.update_one(
        {"id": card_id},
        {"$set": update_fields}
    )
    
    return new_status

# Utility functions
def clean_customer_code(code: str) -> str:
    """Clean customer code by removing fees and extra characters"""
    # Remove everything after comma (fees)
    code = code.split(',')[0].strip()
    # Remove any non-alphanumeric characters except specific ones
    code = re.sub(r'[^\w]', '', code)
    return code.upper()

def map_provider_region(electric_provider: str) -> ProviderRegion:
    """Map electric provider string to enum"""
    mapping = {
        "mien_bac": ProviderRegion.MIEN_BAC,
        "mien_nam": ProviderRegion.MIEN_NAM,
        "hcmc": ProviderRegion.HCMC
    }
    return mapping.get(electric_provider.lower(), ProviderRegion.MIEN_NAM)

def prepare_for_mongo(data):
    """Convert datetime objects to ISO strings for MongoDB storage"""
    if isinstance(data, dict):
        result = {}
        for key, value in data.items():
            if isinstance(value, datetime):
                result[key] = value.isoformat()
            elif isinstance(value, dict):
                result[key] = prepare_for_mongo(value)
            elif isinstance(value, list):
                result[key] = [prepare_for_mongo(item) if isinstance(item, dict) else item for item in value]
            else:
                result[key] = value
        return result
    return data

def parse_from_mongo(item):
    """Parse MongoDB document for JSON serialization"""
    if isinstance(item, dict):
        # Convert MongoDB ObjectId to string for JSON serialization
        if '_id' in item:
            item['id'] = str(item['_id'])
            item.pop('_id', None)
        
        # Parse datetime strings
        for key, value in item.items():
            if isinstance(value, str) and key.endswith('_at'):
                try:
                    item[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                except:
                    pass
            # Handle nested dictionaries
            elif isinstance(value, dict):
                item[key] = parse_from_mongo(value)
            # Handle lists of dictionaries
            elif isinstance(value, list):
                item[key] = [parse_from_mongo(v) if isinstance(v, dict) else v for v in value]
    return item

# External API bill checking function
async def external_check_bill(customer_code: str, provider_region: ProviderRegion) -> CheckBillResult:  
    """Call external bill checking API with timeout"""
    print(
        f"Checking bill for customer code: {customer_code}, "
        f"Provider: {provider_region.value}"
    )
    
    # Use default webhook URL
    webhook_url = "https://n8n.phamthanh.net/webhook/checkbill"

    
    # Mock successful response for PA2204000000
    if customer_code == "PA2204000000":
        await asyncio.sleep(0.1)  # Quick mock delay
        return CheckBillResult(
            customer_code=customer_code,
            full_name="Nguyen Van Test",
            address="123 Test Street",
            amount=150000.0,
            billing_cycle="2024-12",
            status="OK",
            bill_id=str(uuid.uuid4())
        )
    
    # Mock successful response for PA2204000000
    if customer_code == "PA2204000000":
        # Create successful bill record in database
        bill_data = {
            "id": str(uuid.uuid4()),
            "gateway": Gateway.FPT,
            "customer_code": customer_code,
            "provider_region": provider_region,
            "provider_name": provider_region.value,
            "full_name": "Nguyễn Thị Hương",
            "address": "123 Phố Huế, Hai Bà Trưng, Hà Nội",
            "amount": 1850000,
            "billing_cycle": "09/2025",
            "status": BillStatus.AVAILABLE,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        # Save to database
        await db.bills.update_one(
            {"customer_code": customer_code, "gateway": Gateway.FPT},
            {"$set": bill_data},
            upsert=True
        )
        
        return CheckBillResult(
            customer_code=customer_code,
            full_name="Nguyễn Thị Hương",
            address="123 Phố Huế, Hai Bà Trưng, Hà Nội",
            amount=1850000,
            billing_cycle="09/2025",
            status="OK",
            bill_id=bill_data["id"]  # Include bill_id for inventory
        )
    
    # Map provider region to external format
    provider_mapping = {
        ProviderRegion.MIEN_BAC: "mien_bac",
        ProviderRegion.MIEN_NAM: "mien_nam", 
        ProviderRegion.HCMC: "evnhcmc"
    }
    
    external_provider = provider_mapping.get(provider_region, "mien_nam")
    
    # Prepare payload for external webhook
    payload = {
        "bills": [
            {
                "customer_id": customer_code,
                "electric_provider": external_provider,
                "provider_name": external_provider,
                "contractNumber": customer_code,
                "sku": "ELECTRIC_BILL"
            }
        ],
        "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000),
        "request_id": f"fpt_bill_manager_{str(uuid.uuid4())[:8]}",
        "webhookUrl": "https://n8n.phamthanh.net/webhook/checkbill",
        "executionMode": "production"
    }
    
    try:
        # Configure timeout for external API call (30 seconds)
        timeout = aiohttp.ClientTimeout(total=30, connect=10)
        
        async with aiohttp.ClientSession(timeout=timeout) as session:
            print(f"[DEBUG] Making external API call to: https://n8n.phamthanh.net/webhook/checkbill")
            print(f"[DEBUG] Payload: {payload}")
            
            async with session.post(
                "https://n8n.phamthanh.net/webhook/checkbill",
                json=payload,
                headers={"Content-Type": "application/json"}
            ) as response:
                response_text = await response.text()
                print(f"[DEBUG] External API response status: {response.status}")
                print(f"[DEBUG] External API response text: {response_text[:500]}...")
                
                # Parse response - handle different response formats
                try:
                    response_data = json.loads(response_text)
                    
                    # Handle array format (old format)
                    if isinstance(response_data, list) and len(response_data) > 0:
                        first_item = response_data[0]
                        
                        # Check for success in new API format
                        if (first_item.get("status") == 200 and 
                            "data" in first_item and 
                            "bills" in first_item["data"] and 
                            len(first_item["data"]["bills"]) > 0):
                            
                            bill_info = first_item["data"]["bills"][0]
                            
                            # Map external API fields to internal fields
                            full_name = bill_info.get("customerName")
                            address = bill_info.get("address")
                            amount = bill_info.get("moneyAmount")
                            billing_cycle = bill_info.get("month")
                            
                            # Successful bill found
                            bill_data = {
                                "id": str(uuid.uuid4()),
                                "gateway": Gateway.FPT,
                                "customer_code": customer_code,
                                "provider_region": provider_region,
                                "provider_name": external_provider,
                                "full_name": full_name,
                                "address": address,
                                "amount": amount,
                                "billing_cycle": billing_cycle,
                                "status": BillStatus.AVAILABLE,
                                "meta": first_item,
                                "created_at": datetime.now(timezone.utc).isoformat(),
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }
                            
                            # Save to database
                            await db.bills.update_one(
                                {"customer_code": customer_code, "gateway": Gateway.FPT},
                                {"$set": bill_data},
                                upsert=True
                            )
                            
                            return CheckBillResult(
                                customer_code=customer_code,
                                full_name=full_name,
                                address=address,
                                amount=amount,
                                billing_cycle=billing_cycle,
                                status="OK",
                                bill_id=bill_data["id"]
                            )
                        
                        # Check for old format success (has bill data)
                        elif "error" not in first_item and ("customer_code" in str(first_item) or "full_name" in first_item):
                            # Successful bill found (old format)
                            bill_data = {
                                "id": str(uuid.uuid4()),
                                "gateway": Gateway.FPT,
                                "customer_code": customer_code,
                                "provider_region": provider_region,
                                "provider_name": external_provider,
                                "full_name": first_item.get("full_name"),
                                "address": first_item.get("address"),
                                "amount": first_item.get("amount"),
                                "billing_cycle": first_item.get("billing_cycle"),
                                "status": BillStatus.AVAILABLE,
                                "meta": first_item,
                                "created_at": datetime.now(timezone.utc).isoformat(),
                                "updated_at": datetime.now(timezone.utc).isoformat()
                            }
                            
                            # Save to database
                            await db.bills.update_one(
                                {"customer_code": customer_code, "gateway": Gateway.FPT},
                                {"$set": bill_data},
                                upsert=True
                            )
                            
                            return CheckBillResult(
                                customer_code=customer_code,
                                full_name=first_item.get("full_name"),
                                address=first_item.get("address"),
                                amount=first_item.get("amount"),
                                billing_cycle=first_item.get("billing_cycle"),
                                status="OK",
                                bill_id=bill_data["id"]
                            )
                        else:
                            # Error response - extract message from nested error
                            error_message = "Mã không tồn tại"
                            
                            if "error" in first_item and "message" in first_item["error"]:
                                full_message = first_item["error"]["message"]
                                
                                # Method 1: Try to extract Vietnamese error message using regex
                                import re
                                vietnamese_patterns = [
                                    r'"message":"([^"]*(?:không nợ cước)[^"]*)"',
                                    r'"message":"([^"]*(?:không tồn tại)[^"]*)"', 
                                    r'"message":"([^"]*(?:không hợp lệ)[^"]*)"',
                                    r'"message":"([^"]*(?:lỗi)[^"]*)"'
                                ]
                                
                                for pattern in vietnamese_patterns:
                                    match = re.search(pattern, full_message, re.IGNORECASE)
                                    if match:
                                        error_message = match.group(1)
                                        break
                                        
                                # Method 2: If no Vietnamese message found, try JSON parsing
                                if error_message == "Mã không tồn tại":
                                    try:
                                        # Handle escaped JSON string
                                        clean_message = full_message
                                        if '400 - "' in clean_message:
                                            # Extract JSON part after "400 - "
                                            json_start = clean_message.find('400 - "') + 7
                                            json_part = clean_message[json_start:-1]  # Remove trailing quote
                                            json_part = json_part.replace('\\"', '"')  # Unescape quotes
                                            
                                            nested_data = json.loads(json_part)
                                            if "error" in nested_data and "message" in nested_data["error"]:
                                                error_message = nested_data["error"]["message"]
                                    except Exception as e:
                                        logger.debug(f"JSON parsing failed, using regex fallback: {e}")
                                        
                            return CheckBillResult(
                                customer_code=customer_code,
                                status="ERROR",
                                errors={"code": "EXTERNAL_API_ERROR", "message": error_message}
                            )
                    else:
                        return CheckBillResult(
                            customer_code=customer_code,
                            status="ERROR", 
                            errors={"code": "INVALID_RESPONSE", "message": "Phản hồi không hợp lệ từ hệ thống"}
                        )
                        
                except json.JSONDecodeError:
                    return CheckBillResult(
                        customer_code=customer_code,
                        status="ERROR",
                        errors={"code": "PARSE_ERROR", "message": "Không thể phân tích phản hồi"}
                    )
                    
    except asyncio.TimeoutError:
        print(f"[ERROR] External API timeout for customer code: {customer_code}")
        return CheckBillResult(
            customer_code=customer_code,
            status="ERROR",
            errors={"code": "TIMEOUT_ERROR", "message": "Hết thời gian chờ phản hồi từ hệ thống"}
        )
    except aiohttp.ClientError as e:
        print(f"[ERROR] External API client error for {customer_code}: {str(e)}")
        return CheckBillResult(
            customer_code=customer_code,
            status="ERROR", 
            errors={"code": "CONNECTION_ERROR", "message": f"Lỗi kết nối mạng: {str(e)}"}
        )
    except Exception as e:
        logger.error(f"Error calling external webhook for {customer_code}: {str(e)}")
        print(f"[ERROR] Unexpected error for {customer_code}: {str(e)}")
        return CheckBillResult(
            customer_code=customer_code,
            status="ERROR",
            errors={"code": "UNKNOWN_ERROR", "message": f"Lỗi không xác định: {str(e)}"}
        )

# =============================================================================
# AUTHENTICATION API ROUTES
# =============================================================================

@api_router.post("/auth/register", response_model=UserResponse)
async def register_user(user_data: UserCreate):
    """Register a new user"""
    try:
        # Check if username already exists
        existing_user = await db.users.find_one({
            "$or": [
                {"username": user_data.username},
                {"email": user_data.email}
            ]
        })
        
        if existing_user:
            if existing_user.get("username") == user_data.username:
                raise HTTPException(status_code=400, detail="Username already exists")
            else:
                raise HTTPException(status_code=400, detail="Email already exists")
        
        # Check phone if provided
        if user_data.phone:
            phone_exists = await db.users.find_one({"phone": user_data.phone})
            if phone_exists:
                raise HTTPException(status_code=400, detail="Phone number already exists")
        
        # Create new user
        user_dict = {
            "id": str(uuid.uuid4()),
            "username": user_data.username,
            "email": user_data.email,
            "phone": user_data.phone,
            "password": hash_password(user_data.password),
            "full_name": user_data.full_name,
            "role": user_data.role,
            "is_active": True,
            "created_at": datetime.now(timezone.utc),
            "last_login": None
        }
        
        # Insert user
        await db.users.insert_one(user_dict)
        
        # Return user response (without password)
        user_dict.pop("password")
        return UserResponse(**user_dict)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error registering user: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/auth/login", response_model=TokenResponse)
async def login_user(login_data: UserLogin):
    """Login user with username/email/phone and password"""
    try:
        # Find user by username, email, or phone
        user = await db.users.find_one({
            "$or": [
                {"username": login_data.login},
                {"email": login_data.login},
                {"phone": login_data.login}
            ]
        })
        
        if not user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Incorrect username/email/phone or password"
            )
        
        # Verify password
        if not verify_password(login_data.password, user["password"]):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Incorrect username/email/phone or password"  
            )
        
        # Check if user is active
        if not user.get("is_active", True):
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Account is deactivated"
            )
        
        # Update last login
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"last_login": datetime.now(timezone.utc)}}
        )
        
        # Create access token
        access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
        access_token = create_access_token(
            data={"sub": user["id"], "role": user["role"]}, 
            expires_delta=access_token_expires
        )
        
        # Prepare user response (without password)
        user_dict = dict(user)
        user_dict.pop("password")
        user_response = UserResponse(**user_dict)
        
        return TokenResponse(
            access_token=access_token,
            user=user_response
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error during login: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/auth/me", response_model=UserResponse)
async def get_current_user_info(current_user: dict = authenticated_user):
    """Get current user information"""
    user_dict = dict(current_user)
    user_dict.pop("password", None)
    return UserResponse(**user_dict)

@api_router.put("/auth/profile", response_model=UserResponse)
async def update_user_profile(
    profile_data: UserUpdate, 
    current_user: dict = authenticated_user
):
    """Update user profile"""
    try:
        update_data = {}
        
        # Prepare update data
        if profile_data.full_name is not None:
            update_data["full_name"] = profile_data.full_name
        if profile_data.phone is not None:
            # Check if phone already exists for another user
            if profile_data.phone != current_user.get("phone"):
                phone_exists = await db.users.find_one({
                    "phone": profile_data.phone,
                    "id": {"$ne": current_user["id"]}
                })
                if phone_exists:
                    raise HTTPException(status_code=400, detail="Phone number already exists")
            update_data["phone"] = profile_data.phone
        if profile_data.email is not None:
            # Check if email already exists for another user
            if profile_data.email != current_user.get("email"):
                email_exists = await db.users.find_one({
                    "email": profile_data.email,
                    "id": {"$ne": current_user["id"]}
                })
                if email_exists:
                    raise HTTPException(status_code=400, detail="Email already exists")
            update_data["email"] = profile_data.email
        
        if not update_data:
            raise HTTPException(status_code=400, detail="No data provided for update")
        
        update_data["updated_at"] = datetime.now(timezone.utc)
        
        # Update user
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": update_data}
        )
        
        # Get updated user
        updated_user = await db.users.find_one({"id": current_user["id"]})
        updated_user.pop("password", None)
        
        return UserResponse(**updated_user)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating user profile: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/auth/change-password")
async def change_password(
    password_data: PasswordChange,
    current_user: dict = authenticated_user
):
    """Change user password"""
    try:
        # Verify current password
        if not verify_password(password_data.current_password, current_user["password"]):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Current password is incorrect"
            )
        
        # Hash new password
        new_password_hash = hash_password(password_data.new_password)
        
        # Update password
        await db.users.update_one(
            {"id": current_user["id"]},
            {"$set": {
                "password": new_password_hash,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        return {"message": "Password changed successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error changing password: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/auth/users", response_model=List[UserResponse])
async def get_all_users(current_user: dict = manager_or_admin_required):
    """Get all users (Manager/Admin only)"""
    try:
        users_cursor = db.users.find({})
        users = await users_cursor.to_list(None)
        
        # Remove passwords from response
        user_responses = []
        for user in users:
            user.pop("password", None)
            user_responses.append(UserResponse(**user))
            
        return user_responses
        
    except Exception as e:
        logger.error(f"Error getting users: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/auth/users/{user_id}/role")
async def update_user_role(
    user_id: str,
    role_data: dict,
    current_user: dict = admin_required
):
    """Update user role (Admin only)"""
    try:
        new_role = role_data.get("role")
        if new_role not in [role.value for role in UserRole]:
            raise HTTPException(status_code=400, detail="Invalid role")
        
        # Update user role
        result = await db.users.update_one(
            {"id": user_id},
            {"$set": {
                "role": new_role,
                "updated_at": datetime.now(timezone.utc)
            }}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="User not found")
        
        return {"message": "User role updated successfully"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating user role: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# API Routes
@api_router.get("/dashboard/stats", response_model=DashboardStats)
async def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        # Count bills by status
        total_bills = await db.bills.count_documents({})
        available_bills = await db.bills.count_documents({"status": BillStatus.AVAILABLE})
        sold_bills = await db.bills.count_documents({"status": BillStatus.SOLD})
        
        # Count customers
        total_customers = await db.customers.count_documents({})
        
        # Calculate total revenue from sales
        pipeline = [
            {"$group": {"_id": None, "total": {"$sum": "$profit_value"}}}
        ]
        revenue_result = await db.sales.aggregate(pipeline).to_list(1)
        total_revenue = revenue_result[0]["total"] if revenue_result else 0
        
        # Get recent activities (latest 5 sales)
        recent_sales = await db.sales.find().sort("created_at", -1).limit(5).to_list(5)
        recent_activities = []
        
        for sale in recent_sales:
            customer = await db.customers.find_one({"id": sale["customer_id"]})
            customer_name = customer["name"] if customer else "Không xác định"
            
            recent_activities.append({
                "id": sale["id"],
                "type": "sale",
                "description": f"Bán bill cho {customer_name}",
                "amount": sale["total"],
                "profit": sale["profit_value"],
                "created_at": sale["created_at"]
            })
        
        return DashboardStats(
            total_bills=total_bills,
            available_bills=available_bills,
            sold_bills=sold_bills,
            total_customers=total_customers,
            total_revenue=total_revenue,
            recent_activities=recent_activities
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/bill/check", response_model=CheckBillResponse)
async def check_bills(request: CheckBillRequest):
    """Check multiple bills"""
    try:
        print(f"[DEBUG] Starting batch check for {len(request.codes)} bills")
        
        results = []
        ok_count = 0
        error_count = 0
        
        for code in request.codes:
            if not code.strip():
                continue
                
            cleaned_code = clean_customer_code(code)
            result = await external_check_bill(cleaned_code, request.provider_region)
            results.append(result)
            
            if result.status == "OK":
                ok_count += 1
            else:
                error_count += 1
        
        return CheckBillResponse(
            items=results,
            summary={"ok": ok_count, "error": error_count}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/bill/check/single", response_model=CheckBillResult)
async def check_single_bill(
    customer_code: str,
    provider_region: ProviderRegion
):
    """Check single bill for realtime response"""
    try:
        cleaned_code = clean_customer_code(customer_code)
        result = await external_check_bill(cleaned_code, provider_region)
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/bill/debug-payload")
async def debug_bill_payload(
    customer_code: str,
    provider_region: ProviderRegion
):
    """Debug endpoint to show exact payload being sent to external API"""
    # Map provider region to external format
    provider_mapping = {
        ProviderRegion.MIEN_BAC: "mien_bac",
        ProviderRegion.MIEN_NAM: "mien_nam", 
        ProviderRegion.HCMC: "evnhcmc"
    }
    
    external_provider = provider_mapping.get(provider_region, "mien_nam")
    
    # Prepare payload for external webhook
    payload = {
        "bills": [
            {
                "customer_id": customer_code,
                "electric_provider": external_provider,
                "provider_name": external_provider,
                "contractNumber": customer_code,
                "sku": "ELECTRIC_BILL"
            }
        ],
        "timestamp": int(datetime.now(timezone.utc).timestamp() * 1000),
        "request_id": f"fpt_bill_manager_{str(uuid.uuid4())[:8]}",
        "webhookUrl": "https://n8n.phamthanh.net/webhook/checkbill",
        "executionMode": "production"
    }
    
    return {
        "customer_code": customer_code,
        "provider_region": provider_region,
        "external_provider": external_provider,
        "payload": payload,
        "external_api_url": "https://n8n.phamthanh.net/webhook/checkbill"
    }

@api_router.get("/bills", response_model=List[Bill])
async def get_bills(status: Optional[BillStatus] = None, limit: int = 50):
    """Get bills with optional status filter"""
    try:
        query = {}
        if status:
            query["status"] = status
            
        bills = await db.bills.find(query).limit(limit).to_list(limit)
        return [Bill(**parse_from_mongo(bill)) for bill in bills]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/customers/stats", response_model=CustomerStats)
async def get_customer_stats():
    """Get customer statistics"""
    try:
        # Count customers by type and status
        total_customers = await db.customers.count_documents({})
        individual_customers = await db.customers.count_documents({"type": CustomerType.INDIVIDUAL})
        agent_customers = await db.customers.count_documents({"type": CustomerType.AGENT})
        active_customers = await db.customers.count_documents({"is_active": True})
        
        # Calculate total customer value
        pipeline = [
            {"$group": {"_id": None, "total_value": {"$sum": "$total_value"}}}
        ]
        value_result = await db.customers.aggregate(pipeline).to_list(1)
        total_customer_value = value_result[0]["total_value"] if value_result else 0
        
        return CustomerStats(
            total_customers=total_customers,
            individual_customers=individual_customers,
            agent_customers=agent_customers,
            active_customers=active_customers,
            total_customer_value=total_customer_value
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/customers", response_model=List[Customer])
async def get_customers(
    search: Optional[str] = None,
    customer_type: Optional[CustomerType] = None,
    is_active: Optional[bool] = None,
    page: int = 1,
    page_size: int = 20
):
    """Get customers with filtering and pagination"""
    try:
        # Build query
        query = {}
        
        if search:
            query["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"phone": {"$regex": search, "$options": "i"}},
                {"email": {"$regex": search, "$options": "i"}},
                {"address": {"$regex": search, "$options": "i"}}
            ]
        
        if customer_type:
            query["type"] = customer_type
            
        if is_active is not None:
            query["is_active"] = is_active
        
        # Get paginated results
        skip = (page - 1) * page_size
        customers = await db.customers.find(query).skip(skip).limit(page_size).sort("created_at", -1).to_list(page_size)
        
        # Filter out invalid customers and fix old data
        valid_customers = []
        for customer in customers:
            try:
                # Handle old BUSINESS type data
                if customer.get("type") == "BUSINESS":
                    customer["type"] = "AGENT"
                    # Update in database
                    await db.customers.update_one(
                        {"_id": customer["_id"]},
                        {"$set": {"type": "AGENT"}}
                    )
                
                # Ensure all required fields exist with defaults
                customer.setdefault("total_transactions", 0)
                customer.setdefault("total_value", 0.0)
                customer.setdefault("total_bills", 0)
                customer.setdefault("total_cards", 0)
                customer.setdefault("total_profit_generated", 0.0)
                customer.setdefault("is_active", True)
                customer.setdefault("notes", None)
                
                parsed_customer = Customer(**parse_from_mongo(customer))
                valid_customers.append(parsed_customer)
            except Exception as parse_error:
                logger.warning(f"Skipping invalid customer {customer.get('_id', 'unknown')}: {parse_error}")
                continue
        
        return valid_customers
    except Exception as e:
        logger.error(f"Error getting customers: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/customers/export")
async def export_customers_data():
    """Export customers data to Excel"""
    try:
        # Get all customers (simple version)
        customers = await db.customers.find({}).sort("created_at", -1).to_list(None)
        
        # Create Excel file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        workbook = openpyxl.Workbook()
        
        # Customers Sheet
        customers_sheet = workbook.active
        customers_sheet.title = "Khách Hàng"
        
        # Customer headers
        customer_headers = [
            "ID", "Tên khách hàng", "Số điện thoại", "Email", 
            "Địa chỉ", "Loại khách hàng", "Trạng thái", "Tổng giao dịch",
            "Tổng giá trị", "Tổng lợi nhuận", "Ngày tạo", "Ghi chú"
        ]
        
        # Write customer headers
        for col, header in enumerate(customer_headers, 1):
            cell = customers_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Write customer data
        for row, customer in enumerate(customers, 2):
            customers_sheet.cell(row=row, column=1, value=customer.get("id"))
            customers_sheet.cell(row=row, column=2, value=customer.get("name"))
            customers_sheet.cell(row=row, column=3, value=customer.get("phone"))
            customers_sheet.cell(row=row, column=4, value=customer.get("email"))
            customers_sheet.cell(row=row, column=5, value=customer.get("address"))
            customers_sheet.cell(row=row, column=6, value=customer.get("type"))
            customers_sheet.cell(row=row, column=7, value="Hoạt động" if customer.get("is_active") else "Không hoạt động")
            customers_sheet.cell(row=row, column=8, value=customer.get("total_transactions", 0))
            customers_sheet.cell(row=row, column=9, value=customer.get("total_value", 0))
            customers_sheet.cell(row=row, column=10, value=customer.get("total_profit_generated", 0))
            customers_sheet.cell(row=row, column=11, value=customer.get("created_at"))
            customers_sheet.cell(row=row, column=12, value=customer.get("notes"))
        
        # Auto-adjust column widths for customers sheet
        for column in customers_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            customers_sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            BytesIO(excel_buffer.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": "attachment; filename=khach_hang_export.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/customers/{customer_id}", response_model=Customer)
async def get_customer(customer_id: str):
    """Get single customer by ID - supports both UUID and ObjectId lookup"""
    try:
        # Try to find customer by 'id' field first (UUID format)
        customer = await db.customers.find_one({"id": customer_id})
        
        # If not found and customer_id looks like ObjectId, try _id field
        if not customer and len(customer_id) == 24 and all(c in '0123456789abcdef' for c in customer_id.lower()):
            try:
                from bson import ObjectId
                customer = await db.customers.find_one({"_id": ObjectId(customer_id)})
            except:
                pass  # Invalid ObjectId format, continue with None
        
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        return Customer(**parse_from_mongo(customer))
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate):
    """Create new customer"""
    try:
        # Check for duplicate email if provided
        if customer_data.email:
            existing = await db.customers.find_one({"email": customer_data.email})
            if existing:
                raise HTTPException(status_code=400, detail="Email đã tồn tại")
        
        # Create customer
        customer = Customer(
            **customer_data.dict(),
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc)
        )
        
        customer_dict = prepare_for_mongo(customer.dict())
        await db.customers.insert_one(customer_dict)
        
        return customer
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_data: CustomerUpdate):
    """Update customer information - supports both UUID and ObjectId lookup"""
    try:
        # Try to find customer by 'id' field first (UUID format)
        existing = await db.customers.find_one({"id": customer_id})
        
        # If not found and customer_id looks like ObjectId, try _id field
        if not existing and len(customer_id) == 24 and all(c in '0123456789abcdef' for c in customer_id.lower()):
            try:
                from bson import ObjectId
                existing = await db.customers.find_one({"_id": ObjectId(customer_id)})
                # If found by ObjectId, use the actual 'id' field for subsequent operations
                if existing and existing.get('id'):
                    customer_id = existing.get('id')
            except:
                pass  # Invalid ObjectId format, continue with original customer_id
        
        if not existing:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Check for duplicate email if updating email
        if customer_data.email and customer_data.email != existing.get("email"):
            duplicate = await db.customers.find_one({"email": customer_data.email, "id": {"$ne": customer_id}})
            if duplicate:
                raise HTTPException(status_code=400, detail="Email đã tồn tại")
        
        # Update fields
        update_data = {k: v for k, v in customer_data.dict().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.customers.update_one(
            {"id": customer_id},
            {"$set": update_data}
        )
        
        # Return updated customer
        updated_customer = await db.customers.find_one({"id": customer_id})
        return Customer(**parse_from_mongo(updated_customer))
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/inventory/{item_id}")
async def delete_inventory_item(item_id: str):
    """Delete item from inventory"""
    try:
        # Remove from inventory_items collection
        result = await db.inventory_items.delete_one({"id": item_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Không tìm thấy item trong kho")
        
        return {"success": True, "message": "Đã xóa item khỏi kho thành công"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] DAO Exception: {e}")
        print(f"[ERROR] Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/bills/{bill_id}")
async def delete_bill(bill_id: str):
    """Delete bill completely - with validation for sold bills"""
    try:
        # First check if bill exists and get its status
        bill = await db.bills.find_one({"id": bill_id})
        if not bill:
            raise HTTPException(status_code=404, detail="Không tìm thấy bill")
        
        # CRITICAL: Check if bill is already sold or crossed - prevent deletion
        if bill.get("status") == BillStatus.SOLD:
            raise HTTPException(
                status_code=400, 
                detail="Không thể xóa bill đã bán. Bill này đã được tham chiếu trong giao dịch khách hàng."
            )
        
        if bill.get("status") == BillStatus.CROSSED:
            raise HTTPException(
                status_code=400,
                detail="Không thể xóa bill đã gạch. Bill này đã được xác nhận không có nợ cước."
            )
        
        # Check if bill is referenced in any sales (double-check safety)
        sales_using_bill = await db.sales.find_one({"bill_ids": bill_id})
        if sales_using_bill:
            raise HTTPException(
                status_code=400,
                detail="Không thể xóa bill đã có giao dịch. Bill này đang được tham chiếu trong lịch sử bán hàng."
            )
        
        # Safe to delete - bill is available/pending
        result = await db.bills.delete_one({"id": bill_id})
        
        # Also remove from inventory if exists
        await db.inventory_items.delete_many({"bill_id": bill_id})
        
        return {"success": True, "message": "Đã xóa bill thành công"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] DAO Exception: {e}")
        print(f"[ERROR] Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/bills/{bill_id}")
async def get_bill(bill_id: str):
    """Get single bill by ID - used for existence check"""
    try:
        bill = await db.bills.find_one({"id": bill_id})
        if not bill:
            raise HTTPException(status_code=404, detail="Không tìm thấy bill")
        
        return Bill(**parse_from_mongo(bill))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting bill {bill_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/bills/{bill_id}")
async def update_bill(bill_id: str, bill_data: dict):
    """Update bill information - used for recheck functionality"""
    try:
        print(f"PUT /bills/{bill_id} called with data keys: {list(bill_data.keys())}")
        
        # Check if bill exists
        existing_bill = await db.bills.find_one({"id": bill_id})
        if not existing_bill:
            print(f"Bill {bill_id} not found in database")
            raise HTTPException(status_code=404, detail="Không tìm thấy bill")
        
        print(f"Found existing bill: {existing_bill.get('id')}")
        
        # Prepare update data - only update fields that are provided and allowed
        update_data = {
            "updated_at": datetime.now(timezone.utc)
        }
        
        # Update allowed fields if provided
        allowed_fields = [
            "customer_code", "provider_region", "full_name", "address", 
            "amount", "billing_cycle", "status", "raw_status", 
            "error_code", "error_message", "meta", "note", "last_checked"
        ]
        
        for field in allowed_fields:
            if field in bill_data and bill_data[field] is not None:
                if field == "last_checked" and isinstance(bill_data[field], str):
                    # Parse ISO string to datetime
                    try:
                        update_data[field] = datetime.fromisoformat(bill_data[field].replace('Z', '+00:00'))
                    except:
                        update_data[field] = datetime.now(timezone.utc)
                else:
                    update_data[field] = bill_data[field]
        
        print(f"Update data prepared with fields: {list(update_data.keys())}")
        
        # Update bill in database
        result = await db.bills.update_one(
            {"id": bill_id},
            {"$set": update_data}
        )
        
        print(f"Update result: matched={result.matched_count}, modified={result.modified_count}")
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Không tìm thấy bill để cập nhật")
        
        # Get updated bill
        updated_bill = await db.bills.find_one({"id": bill_id})
        print(f"Retrieved updated bill successfully")
        
        return {
            "id": updated_bill["id"],
            "customer_code": updated_bill["customer_code"],
            "status": updated_bill["status"],
            "last_checked": updated_bill.get("last_checked"),
            "updated_at": updated_bill["updated_at"]
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error updating bill: {str(e)}")
        logger.error(f"Error updating bill {bill_id}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Lỗi cập nhật bill: {str(e)}")

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str):
    """Delete customer and all related data - supports both UUID and ObjectId lookup"""
    try:
        # Try to find customer by 'id' field first (UUID format)
        customer = await db.customers.find_one({"id": customer_id})
        actual_customer_id = customer_id
        
        # If not found and customer_id looks like ObjectId, try _id field
        if not customer and len(customer_id) == 24 and all(c in '0123456789abcdef' for c in customer_id.lower()):
            try:
                from bson import ObjectId
                customer = await db.customers.find_one({"_id": ObjectId(customer_id)})
                # If found by ObjectId, use the actual 'id' field for cascade deletions
                if customer and customer.get('id'):
                    actual_customer_id = customer.get('id')
            except:
                pass  # Invalid ObjectId format, continue with original customer_id
        
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Get all sales for this customer to find related bills
        sales = await db.sales.find({"customer_id": actual_customer_id}).to_list(None)
        
        deleted_stats = {
            "customer": 0,
            "transactions": 0, 
            "bills": 0,
            "inventory_items": 0,
            "credit_cards": 0,
            "credit_card_transactions": 0
        }
        
        # Delete credit cards and their transactions first
        credit_cards = await db.credit_cards.find({"customer_id": actual_customer_id}).to_list(None)
        if credit_cards:
            card_ids = [card["id"] for card in credit_cards]
            
            # Delete credit card transactions
            cc_tx_delete_result = await db.credit_card_transactions.delete_many({"card_id": {"$in": card_ids}})
            deleted_stats["credit_card_transactions"] = cc_tx_delete_result.deleted_count
            
            # Delete credit cards
            cc_delete_result = await db.credit_cards.delete_many({"customer_id": actual_customer_id})
            deleted_stats["credit_cards"] = cc_delete_result.deleted_count
        
        # Delete related data in order
        if sales:
            # Get all bill IDs from sales
            bill_ids = []
            for sale in sales:
                if sale.get("bill_ids"):
                    bill_ids.extend(sale["bill_ids"])
            
            # Remove duplicate bill IDs
            unique_bill_ids = list(set(bill_ids))
            
            # Delete inventory items for these bills
            inventory_delete_result = await db.inventory_items.delete_many({"bill_id": {"$in": unique_bill_ids}})
            deleted_stats["inventory_items"] = inventory_delete_result.deleted_count
            
            # Delete the bills themselves
            bills_delete_result = await db.bills.delete_many({"id": {"$in": unique_bill_ids}})
            deleted_stats["bills"] = bills_delete_result.deleted_count
            
            # Delete all sales/transactions for this customer
            sales_delete_result = await db.sales.delete_many({"customer_id": actual_customer_id})
            deleted_stats["transactions"] = sales_delete_result.deleted_count
        
        # Finally delete the customer
        customer_delete_result = await db.customers.delete_one({"id": actual_customer_id})
        deleted_stats["customer"] = customer_delete_result.deleted_count
        
        if customer_delete_result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Không thể xóa khách hàng")
        
        # Create detailed message
        message_parts = [f"Đã xóa khách hàng '{customer.get('name', 'N/A')}' thành công"]
        if deleted_stats["transactions"] > 0:
            message_parts.append(f"{deleted_stats['transactions']} giao dịch")
        if deleted_stats["bills"] > 0:
            message_parts.append(f"{deleted_stats['bills']} bills")  
        if deleted_stats["inventory_items"] > 0:
            message_parts.append(f"{deleted_stats['inventory_items']} items khỏi kho")
        if deleted_stats["credit_cards"] > 0:
            message_parts.append(f"{deleted_stats['credit_cards']} thẻ tín dụng")
        if deleted_stats["credit_card_transactions"] > 0:
            message_parts.append(f"{deleted_stats['credit_card_transactions']} giao dịch thẻ")
            
        return {
            "success": True, 
            "message": " và ".join(message_parts),
            "deleted_stats": deleted_stats
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/customers/{customer_id}/transactions")
async def get_customer_transactions(customer_id: str, limit: int = 50, offset: int = 0):
    """Get customer's transaction history - supports both UUID and ObjectId lookup"""
    try:
        # Try to find customer by 'id' field first (UUID format)
        customer = await db.customers.find_one({"id": customer_id})
        actual_customer_id = customer_id
        
        # If not found and customer_id looks like ObjectId, try _id field
        if not customer and len(customer_id) == 24 and all(c in '0123456789abcdef' for c in customer_id.lower()):
            try:
                from bson import ObjectId
                customer = await db.customers.find_one({"_id": ObjectId(customer_id)})
                # If found by ObjectId, use the actual 'id' field for transaction queries
                if customer and customer.get('id'):
                    actual_customer_id = customer.get('id')
            except:
                pass  # Invalid ObjectId format, continue with original customer_id
        
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Get sales for this customer
        sales = await db.sales.find({"customer_id": actual_customer_id}).sort("created_at", -1).to_list(100)
        
        transactions = []
        for sale in sales:
            if sale.get("transaction_type") == "ELECTRIC_BILL":
                # Handle electric bill transactions
                bill_codes = []
                if sale.get("bill_ids"):
                    bills = await db.bills.find({"id": {"$in": sale["bill_ids"]}}).to_list(None)
                    bill_codes = [bill.get("customer_code", "") for bill in bills]
                
                transactions.append({
                    "id": sale["id"],
                    "type": "SALE",
                    "total": sale["total"],
                    "profit_value": sale["profit_value"],
                    "payback": sale["payback"],
                    "method": sale["method"],
                    "status": sale["status"],
                    "notes": sale.get("notes"),
                    "bill_codes": bill_codes,
                    "created_at": sale["created_at"]
                })
                
            elif sale.get("transaction_type") == "CREDIT_CARD":
                # Handle credit card transactions - extract card number from notes
                bill_codes = []
                notes = sale.get("notes", "")
                # Extract ****1234 pattern from notes (alphanumeric)
                import re
                card_match = re.search(r'\*{4}([A-Z0-9]{4})', notes)
                if card_match:
                    bill_codes = [f"****{card_match.group(1)}"]
                
                transactions.append({
                    "id": sale["id"],
                    "type": "SALE",
                    "total": sale["total"],
                    "profit_value": sale["profit_value"],
                    "payback": sale["payback"],
                    "method": sale["method"],
                    "status": sale["status"],
                    "notes": sale.get("notes"),
                    "bill_codes": bill_codes,  # Will show ****1234 for credit cards
                    "created_at": sale["created_at"]
                })
        
        return {
            "customer": Customer(**parse_from_mongo(customer)),
            "transactions": transactions,
            "summary": {
                "total_transactions": len(transactions),
                "total_value": sum(t["total"] for t in transactions),
                "total_profit": sum(t["profit_value"] for t in transactions)
            }
        }
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

# Inventory/Kho Bill APIs
@api_router.get("/inventory/stats", response_model=InventoryStats)
async def get_inventory_stats():
    """Get inventory statistics"""
    try:
        # Count bills in inventory
        inventory_bills = await db.inventory_items.count_documents({})
        
        # Count ALL bills in bills collection for "Tất cả bills" tab
        total_bills_in_system = await db.bills.count_documents({})
        
        # Count bills in inventory by their bill status
        pipeline = [
            {
                "$lookup": {
                    "from": "bills",
                    "localField": "bill_id", 
                    "foreignField": "id",
                    "as": "bill_info"
                }
            },
            {
                "$unwind": "$bill_info"
            },
            {
                "$group": {
                    "_id": "$bill_info.status",
                    "count": {"$sum": 1},
                    "total_value": {"$sum": {"$ifNull": ["$bill_info.amount", 0]}}
                }
            }
        ]
        
        status_counts = await db.inventory_items.aggregate(pipeline).to_list(10)
        
        available_bills = 0
        pending_bills = 0
        sold_bills = 0
        total_value = 0
        
        for item in status_counts:
            status = item["_id"]
            count = item["count"]
            value = item["total_value"]
            
            if status == BillStatus.AVAILABLE:
                available_bills = count
            elif status == BillStatus.PENDING:
                pending_bills = count
            elif status == BillStatus.SOLD:
                sold_bills = count
            
            total_value += value
        
        return InventoryStats(
            total_bills=inventory_bills,
            available_bills=available_bills,
            pending_bills=pending_bills,
            sold_bills=sold_bills,
            total_value=total_value,
            total_bills_in_system=total_bills_in_system
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/inventory", response_model=List[InventoryResponse])
async def get_inventory_items(
    status: Optional[BillStatus] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 20
):
    """Get inventory items with filtering and pagination"""
    try:
        # Build aggregation pipeline
        pipeline = [
            {
                "$lookup": {
                    "from": "bills",
                    "localField": "bill_id",
                    "foreignField": "id", 
                    "as": "bill_info"
                }
            },
            {
                "$unwind": "$bill_info"
            }
        ]
        
        # Add filters
        match_conditions = {}
        
        if status:
            match_conditions["bill_info.status"] = status
            
        if search:
            match_conditions["$or"] = [
                {"bill_info.customer_code": {"$regex": search, "$options": "i"}},
                {"bill_info.full_name": {"$regex": search, "$options": "i"}},
                {"bill_info.address": {"$regex": search, "$options": "i"}}
            ]
        
        if match_conditions:
            pipeline.append({"$match": match_conditions})
        
        # Add sorting and pagination
        pipeline.extend([
            {"$sort": {"created_at": -1}},
            {"$skip": (page - 1) * page_size},
            {"$limit": page_size}
        ])
        
        inventory_items = await db.inventory_items.aggregate(pipeline).to_list(page_size)
        
        # Convert to response model
        results = []
        for item in inventory_items:
            bill_info = item["bill_info"]
            results.append(InventoryResponse(
                id=item["id"],
                bill_id=item["bill_id"],
                customer_code=bill_info["customer_code"],
                full_name=bill_info.get("full_name"),
                address=bill_info.get("address"),
                amount=bill_info.get("amount"),
                billing_cycle=bill_info.get("billing_cycle"),
                provider_region=bill_info["provider_region"],
                status=bill_info["status"],
                note=item.get("note"),
                batch_id=item.get("batch_id"),
                created_at=datetime.fromisoformat(item["created_at"].replace('Z', '+00:00')) if isinstance(item["created_at"], str) else item["created_at"]
            ))
        
        return results
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/inventory/add")
async def add_to_inventory(request: AddToInventoryRequest):
    """Add bills to inventory"""
    try:
        # Generate batch ID if batch name provided
        batch_id = None
        if request.batch_name:
            batch_id = f"batch_{str(uuid.uuid4())[:8]}_{request.batch_name.replace(' ', '_')}"
        
        added_items = []
        successful_adds = 0
        
        for bill_id in request.bill_ids:
            # Check if bill exists and is available
            bill = await db.bills.find_one({"id": bill_id})
            if not bill:
                # Try to find by customer_code if bill_id lookup fails
                bill = await db.bills.find_one({"customer_code": bill_id, "status": BillStatus.AVAILABLE})
                if not bill:
                    continue
                bill_id = bill["id"]  # Use actual bill ID
                
            if bill["status"] != BillStatus.AVAILABLE:
                continue
            
            # Check if already in inventory
            existing = await db.inventory_items.find_one({"bill_id": bill_id})
            if existing:
                continue
            
            # Create inventory item
            inventory_item = {
                "id": str(uuid.uuid4()),
                "bill_id": bill_id,
                "note": request.note,
                "batch_id": batch_id,
                "added_by": "system",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.inventory_items.insert_one(inventory_item)
            added_items.append(inventory_item)
            successful_adds += 1
        
        return {
            "success": True,
            "added_count": successful_adds,
            "batch_id": batch_id,
            "message": f"Đã thêm {successful_adds} bill vào kho thành công"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/bills/create", response_model=Bill)
async def create_bill_manual(bill_data: BillCreate):
    """Create new bill manually and add to inventory"""
    try:
        # Check if bill with same customer_code already exists
        existing_bill = await db.bills.find_one({
            "customer_code": bill_data.customer_code,
            "provider_region": bill_data.provider_region
        })
        
        if existing_bill:
            raise HTTPException(
                status_code=400, 
                detail=f"Bill với mã điện {bill_data.customer_code} đã tồn tại"
            )
        
        # Create new bill
        bill = Bill(
            gateway=Gateway.FPT,  # Default to FPT
            customer_code=bill_data.customer_code,
            provider_region=bill_data.provider_region,
            provider_name=bill_data.provider_region.value.lower(),
            full_name=bill_data.full_name,
            address=bill_data.address,
            amount=bill_data.amount,
            billing_cycle=bill_data.billing_cycle,
            status=bill_data.status
        )
        
        # Convert to dict for MongoDB
        bill_dict = bill.dict()
        bill_dict["created_at"] = bill.created_at.isoformat()
        bill_dict["updated_at"] = bill.updated_at.isoformat()
        
        # Save to database
        await db.bills.insert_one(bill_dict)
        
        # If status is AVAILABLE, automatically add to inventory
        if bill_data.status == BillStatus.AVAILABLE:
            inventory_item = {
                "id": str(uuid.uuid4()),
                "bill_id": bill.id,
                "note": "Được thêm thủ công",
                "batch_id": None,
                "added_by": "manual",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.inventory_items.insert_one(inventory_item)
        
        return bill
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] DAO Exception: {e}")
        print(f"[ERROR] Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Import/Export APIs
@api_router.post("/inventory/import/preview")
async def preview_import_data(file: UploadFile = File(...)):
    """Preview imported Excel data before saving"""
    try:
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File phải có định dạng Excel (.xlsx hoặc .xls)")
        
        # Read Excel file
        import openpyxl
        import io
        
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        worksheet = workbook.active
        
        # Expected columns: Mã điện, Nhà cung cấp, Tên khách hàng, Địa chỉ, Nợ cước, Chu kỳ thanh toán, Trạng thái
        expected_headers = ["Mã điện", "Nhà cung cấp", "Tên khách hàng", "Địa chỉ", "Nợ cước", "Chu kỳ thanh toán", "Trạng thái"]
        
        # Read header row
        headers = [cell.value for cell in worksheet[1]]
        
        # Validate headers
        if not all(header in headers for header in ["Mã điện"]):  # Only Mã điện is required
            raise HTTPException(status_code=400, detail="File Excel thiếu cột bắt buộc 'Mã điện'")
        
        # Read data rows
        preview_data = []
        errors = []
        
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
                
            try:
                # Map columns to data
                row_data = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_data[header] = row[i]
                
                # Validate required fields
                if not row_data.get("Mã điện"):
                    errors.append(f"Dòng {row_num}: Thiếu mã điện")
                    continue
                
                # Map provider region
                provider_map = {
                    "Miền Bắc": "MIEN_BAC",
                    "Miền Nam": "MIEN_NAM", 
                    "TP.HCM": "HCMC",
                    "HCMC": "HCMC"
                }
                
                provider = row_data.get("Nhà cung cấp", "Miền Nam")
                mapped_provider = provider_map.get(provider, "MIEN_NAM")
                
                # Format data
                preview_item = {
                    "customer_code": str(row_data.get("Mã điện", "")).strip(),
                    "provider_region": mapped_provider,
                    "full_name": str(row_data.get("Tên khách hàng", "")).strip() or None,
                    "address": str(row_data.get("Địa chỉ", "")).strip() or None,
                    "amount": float(row_data.get("Nợ cước", 0)) if row_data.get("Nợ cước") else None,
                    "billing_cycle": str(row_data.get("Chu kỳ thanh toán", "")).strip() or None,
                    "status": "AVAILABLE",  # Default status
                    "row_number": row_num
                }
                
                preview_data.append(preview_item)
                
            except Exception as e:
                errors.append(f"Dòng {row_num}: Lỗi xử lý dữ liệu - {str(e)}")
        
        return {
            "success": True,
            "total_rows": len(preview_data),
            "errors": errors,
            "data": preview_data[:50],  # Preview first 50 rows
            "has_more": len(preview_data) > 50
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Lỗi xử lý file: {str(e)}")

@api_router.post("/inventory/import/confirm")
async def confirm_import_data(import_data: Dict[str, Any]):
    """Confirm and save imported data"""
    try:
        bills_data = import_data.get("data", [])
        successful_imports = 0
        errors = []
        
        for item in bills_data:
            try:
                # Check if bill already exists
                existing_bill = await db.bills.find_one({
                    "customer_code": item["customer_code"],
                    "provider_region": item["provider_region"]
                })
                
                if existing_bill:
                    errors.append(f"Mã điện {item['customer_code']} đã tồn tại")
                    continue
                
                # Create new bill
                bill = Bill(
                    gateway=Gateway.FPT,
                    customer_code=item["customer_code"],
                    provider_region=ProviderRegion(item["provider_region"]),
                    provider_name=item["provider_region"].lower(),
                    full_name=item.get("full_name"),
                    address=item.get("address"),
                    amount=item.get("amount"),
                    billing_cycle=item.get("billing_cycle"),
                    status=BillStatus.AVAILABLE
                )
                
                # Save to database
                bill_dict = bill.dict()
                bill_dict["created_at"] = bill.created_at.isoformat()
                bill_dict["updated_at"] = bill.updated_at.isoformat()
                await db.bills.insert_one(bill_dict)
                
                # Add to inventory
                inventory_item = {
                    "id": str(uuid.uuid4()),
                    "bill_id": bill.id,
                    "note": "Import từ Excel",
                    "batch_id": None,
                    "added_by": "import",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.inventory_items.insert_one(inventory_item)
                
                successful_imports += 1
                
            except Exception as e:
                errors.append(f"Mã điện {item.get('customer_code', 'unknown')}: {str(e)}")
        
        return {
            "success": True,
            "imported_count": successful_imports,
            "total_count": len(bills_data),
            "errors": errors,
            "message": f"Đã import thành công {successful_imports}/{len(bills_data)} bills"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/inventory/export")
async def export_inventory_data(
    status: Optional[BillStatus] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    provider_region: Optional[ProviderRegion] = None
):
    """Export inventory data to Excel"""
    try:
        # Build query
        query = {}
        
        # Build bills query
        bills_query = {}
        if status:
            bills_query["status"] = status
        if provider_region:
            bills_query["provider_region"] = provider_region
            
        # Date filter
        if start_date or end_date:
            date_filter = {}
            if start_date:
                date_filter["$gte"] = start_date
            if end_date:
                date_filter["$lte"] = end_date
            bills_query["created_at"] = date_filter
        
        # Get inventory items with bills
        pipeline = [
            {
                "$lookup": {
                    "from": "bills",
                    "localField": "bill_id",
                    "foreignField": "id",
                    "as": "bill"
                }
            },
            {
                "$unwind": "$bill"
            }
        ]
        
        # Add match stage if filters exist
        if bills_query:
            pipeline.append({"$match": {f"bill.{k}": v for k, v in bills_query.items()}})
        
        pipeline.append({"$sort": {"created_at": -1}})
        
        inventory_items = await db.inventory_items.aggregate(pipeline).to_list(None)
        
        # Create Excel file
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from io import BytesIO
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Kho Bill"
        
        # Headers
        headers = [
            "Mã điện", "Tên khách hàng", "Địa chỉ", "Nợ cước", 
            "Chu kỳ thanh toán", "Nhà cung cấp", "Trạng thái",
            "Ngày thêm", "Ghi chú"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, item in enumerate(inventory_items, 2):
            bill = item["bill"]
            worksheet.cell(row=row, column=1, value=bill.get("customer_code"))
            worksheet.cell(row=row, column=2, value=bill.get("full_name"))
            worksheet.cell(row=row, column=3, value=bill.get("address"))
            worksheet.cell(row=row, column=4, value=bill.get("amount"))
            worksheet.cell(row=row, column=5, value=bill.get("billing_cycle"))
            worksheet.cell(row=row, column=6, value=bill.get("provider_name"))
            worksheet.cell(row=row, column=7, value=bill.get("status"))
            worksheet.cell(row=row, column=8, value=item.get("created_at"))
            worksheet.cell(row=row, column=9, value=item.get("note"))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Return file
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            BytesIO(excel_buffer.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": "attachment; filename=kho_bill_export.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/inventory/template")
async def download_import_template():
    """Download Excel template for importing bills"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Template Import Bills"
        
        # Headers
        headers = [
            "Mã điện", "Nhà cung cấp", "Tên khách hàng", 
            "Địa chỉ", "Nợ cước", "Chu kỳ thanh toán", "Trạng thái"
        ]
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Add sample data
        sample_data = [
            ["PA2024000001", "Miền Nam", "Nguyễn Văn A", "123 Đường ABC, Quận 1", 850000, "10/2025", "Có Sẵn"],
            ["PA2024000002", "Miền Bắc", "Trần Thị B", "456 Đường DEF, Quận 2", 920000, "10/2025", "Có Sẵn"],
            ["PA2024000003", "TP.HCM", "Lê Văn C", "789 Đường GHI, Quận 3", 750000, "11/2025", "Có Sẵn"]
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                worksheet.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        # Add instructions
        instructions = [
            "",
            "HƯỚNG DẪN SỬ DỤNG:",
            "1. Mã điện: Bắt buộc phải có",
            "2. Nhà cung cấp: Miền Bắc / Miền Nam / TP.HCM",
            "3. Nợ cước: Nhập số tiền (VD: 850000)",
            "4. Chu kỳ thanh toán: Định dạng MM/YYYY (VD: 10/2025)",
            "5. Trạng thái: Có Sẵn / Chờ Xử Lý / Đã Bán",
            "6. Xóa các dòng mẫu trước khi import dữ liệu thật"
        ]
        
        for i, instruction in enumerate(instructions, len(sample_data) + 3):
            worksheet.cell(row=i, column=1, value=instruction)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            BytesIO(excel_buffer.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": "attachment; filename=template_import_bills.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Credit Card APIs
@api_router.get("/credit-cards/stats", response_model=CreditCardStats)
async def get_credit_card_stats():
    """Get credit card statistics"""
    try:
        total_cards = await db.credit_cards.count_documents({})
        paid_off_cards = await db.credit_cards.count_documents({"status": CardStatus.PAID_OFF})
        need_payment_cards = await db.credit_cards.count_documents({"status": CardStatus.NEED_PAYMENT})
        not_due_cards = await db.credit_cards.count_documents({"status": CardStatus.NOT_DUE})
        
        # Calculate total credit limit
        pipeline = [
            {"$group": {"_id": None, "total_limit": {"$sum": "$credit_limit"}}}
        ]
        limit_result = await db.credit_cards.aggregate(pipeline).to_list(1)
        total_credit_limit = limit_result[0]["total_limit"] if limit_result else 0.0
        
        return CreditCardStats(
            total_cards=total_cards,
            paid_off_cards=paid_off_cards,
            need_payment_cards=need_payment_cards,
            not_due_cards=not_due_cards,
            total_credit_limit=total_credit_limit
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/credit-cards", response_model=List[CreditCard])
async def get_credit_cards(
    customer_id: Optional[str] = None,
    status: Optional[CardStatus] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 20
):
    """Get credit cards with filtering and pagination (with real-time status)"""
    try:
        # Build query
        query = {}
        
        if customer_id:
            query["customer_id"] = customer_id
            
        if search:
            query["$or"] = [
                {"customer_name": {"$regex": search, "$options": "i"}},
                {"cardholder_name": {"$regex": search, "$options": "i"}},
                {"bank_name": {"$regex": search, "$options": "i"}},
                {"card_number": {"$regex": search, "$options": "i"}}
            ]
        
        # Calculate skip
        skip = (page - 1) * page_size
        
        # Get credit cards
        cards = await db.credit_cards.find(query).skip(skip).limit(page_size).sort("created_at", -1).to_list(page_size)
        
        # Update real-time status for each card
        current_date = datetime.now(timezone.utc)
        updated_cards = []
        
        for card in cards:
            # Calculate real-time status
            real_time_status = calculate_card_status_realtime(card, current_date)
            
            # Update status if changed
            if card.get("status") != real_time_status.value:
                await update_card_cycle_status(card["id"], current_date)
                card["status"] = real_time_status.value
            
            updated_cards.append(card)
        
        # Apply status filter after real-time calculation
        if status:
            updated_cards = [card for card in updated_cards if card.get("status") == status.value]
        
        return [CreditCard(**parse_from_mongo(card)) for card in updated_cards]
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/credit-cards", response_model=CreditCard)
async def create_credit_card(card_data: CreditCardCreate):
    """Create new credit card"""
    try:
        # Validate customer exists
        customer = await db.customers.find_one({"id": card_data.customer_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Check for duplicate card number
        existing_card = await db.credit_cards.find_one({"card_number": card_data.card_number})
        if existing_card:
            raise HTTPException(status_code=400, detail="Số thẻ đã tồn tại")
        
        # Create card
        card = CreditCard(
            customer_id=card_data.customer_id,
            customer_name=customer["name"],
            card_number=card_data.card_number,
            cardholder_name=card_data.cardholder_name,
            bank_name=card_data.bank_name,
            card_type=card_data.card_type,
            expiry_date=card_data.expiry_date,
            ccv=card_data.ccv,
            statement_date=card_data.statement_date,
            payment_due_date=card_data.payment_due_date,
            credit_limit=card_data.credit_limit,
            status=card_data.status,
            notes=card_data.notes
        )
        
        # Save to database
        card_dict = prepare_for_mongo(card.dict())
        await db.credit_cards.insert_one(card_dict)
        
        # Update customer total_cards count
        await db.customers.update_one(
            {"id": card_data.customer_id},
            {"$inc": {"total_cards": 1}}
        )
        
        # Log activity
        await log_activity(ActivityCreate(
            type=ActivityType.CARD_CREATE,
            title=f"Thêm thẻ ****{card_data.card_number[-4:]}",
            description=f"Thẻ {card_data.card_type.value} - {card_data.bank_name}",
            customer_id=card_data.customer_id,
            customer_name=customer["name"],
            amount=card_data.credit_limit,
            status="SUCCESS",
            metadata={
                "card_type": card_data.card_type.value,
                "bank_name": card_data.bank_name,
                "credit_limit": card_data.credit_limit
            }
        ))
        
        return card
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/credit-cards/{card_id}")
async def update_credit_card(card_id: str, card_data: CreditCardUpdate):
    """Update credit card - supports both UUID and ObjectId lookup"""
    try:
        # Try to find credit card by 'id' field first (UUID format)
        existing_card = await db.credit_cards.find_one({"id": card_id})
        actual_card_id = card_id
        
        # If not found and card_id looks like ObjectId, try _id field
        if not existing_card and len(card_id) == 24 and all(c in '0123456789abcdef' for c in card_id.lower()):
            try:
                from bson import ObjectId
                existing_card = await db.credit_cards.find_one({"_id": ObjectId(card_id)})
                # If found by ObjectId, use the actual 'id' field for subsequent operations
                if existing_card and existing_card.get('id'):
                    actual_card_id = existing_card.get('id')
            except:
                pass  # Invalid ObjectId format, continue with original card_id
        
        if not existing_card:
            raise HTTPException(status_code=404, detail="Không tìm thấy thẻ tín dụng")
        
        # Prepare update data
        update_data = {}
        for field, value in card_data.dict(exclude_unset=True).items():
            if value is not None:
                if isinstance(value, Enum):
                    update_data[field] = value.value
                else:
                    update_data[field] = value
        
        if update_data:
            update_data["updated_at"] = datetime.now(timezone.utc)
            
            # Update card
            await db.credit_cards.update_one(
                {"id": card_id},
                {"$set": prepare_for_mongo(update_data)}
            )
        
        # Get updated card
        updated_card = await db.credit_cards.find_one({"id": card_id})
        return CreditCard(**parse_from_mongo(updated_card))
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/credit-cards/{card_id}")
async def delete_credit_card(card_id: str):
    """Delete credit card and all related transactions - supports both UUID and ObjectId lookup"""
    try:
        # Try to find credit card by 'id' field first (UUID format)
        credit_card = await db.credit_cards.find_one({"id": card_id})
        actual_card_id = card_id
        
        # If not found and card_id looks like ObjectId, try _id field
        if not credit_card and len(card_id) == 24 and all(c in '0123456789abcdef' for c in card_id.lower()):
            try:
                from bson import ObjectId
                credit_card = await db.credit_cards.find_one({"_id": ObjectId(card_id)})
                # If found by ObjectId, use the actual 'id' field for cascade deletions
                if credit_card and credit_card.get('id'):
                    actual_card_id = credit_card.get('id')
            except:
                pass  # Invalid ObjectId format, continue with original card_id
        
        if not credit_card:
            raise HTTPException(status_code=404, detail="Không tìm thấy thẻ tín dụng")
        
        # Check for existing transactions - WARNING but allow deletion
        transaction_count = await db.credit_card_transactions.count_documents({"card_id": actual_card_id})
        sales_count = await db.sales.count_documents({"notes": {"$regex": f"****{credit_card['card_number'][-4:]}"}})
        
        # Delete card (transactions will be preserved for reporting)
        result = await db.credit_cards.delete_one({"id": actual_card_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Không tìm thấy thẻ tín dụng")
        
        # Update customer total_cards count
        await db.customers.update_one(
            {"id": credit_card["customer_id"]},
            {"$inc": {"total_cards": -1}}
        )
        
        # Return success message with transaction preservation info
        message = f"Đã xóa thẻ thành công"
        if transaction_count > 0 or sales_count > 0:
            message += f". Đã giữ lại {transaction_count} giao dịch đáo thẻ và {sales_count} giao dịch khách hàng để báo cáo doanh thu."
        
        return {"success": True, "message": message}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] DAO Exception: {e}")
        print(f"[ERROR] Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/credit-cards/{card_id}/detail")
async def get_credit_card_detail(card_id: str):
    """Get detailed credit card information - supports both UUID and ObjectId lookup"""
    try:
        # Try to find credit card by 'id' field first (UUID format)
        card = await db.credit_cards.find_one({"id": card_id})
        
        # If not found and card_id looks like ObjectId, try _id field
        if not card and len(card_id) == 24 and all(c in '0123456789abcdef' for c in card_id.lower()):
            try:
                from bson import ObjectId
                card = await db.credit_cards.find_one({"_id": ObjectId(card_id)})
                # If found by ObjectId, use the actual 'id' field for subsequent queries
                if card and card.get('id'):
                    card_id = card.get('id')
            except:
                pass  # Invalid ObjectId format, continue with original card_id
        
        if not card:
            raise HTTPException(status_code=404, detail="Không tìm thấy thẻ tín dụng")
        
        # Get customer info
        customer = await db.customers.find_one({"id": card["customer_id"]})
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Get recent transactions (latest 3)
        transactions = await db.credit_card_transactions.find(
            {"card_id": card_id}
        ).sort("created_at", -1).limit(3).to_list(3)
        
        # Get total transaction count
        total_transactions = await db.credit_card_transactions.count_documents({"card_id": card_id})
        
        return {
            "card": CreditCard(**parse_from_mongo(card)),
            "customer": {
                "id": customer["id"],
                "name": customer["name"],
                "phone": customer.get("phone"),
                "type": customer["type"]
            },
            "recent_transactions": [CreditCardTransaction(**parse_from_mongo(t)) for t in transactions],
            "total_transactions": total_transactions
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] DAO Exception: {e}")
        print(f"[ERROR] Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Credit Card Transaction APIs
@api_router.get("/credit-cards/{card_id}/transactions")
async def get_card_transactions(card_id: str, page: int = 1, page_size: int = 3):
    """Get credit card transactions with pagination"""
    try:
        # Verify card exists
        card = await db.credit_cards.find_one({"id": card_id})
        if not card:
            raise HTTPException(status_code=404, detail="Không tìm thấy thẻ")
        
        # Get transactions for this card
        skip = (page - 1) * page_size
        transactions = await db.credit_card_transactions.find(
            {"card_id": card_id}
        ).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
        
        # Get total count
        total_count = await db.credit_card_transactions.count_documents({"card_id": card_id})
        
        return {
            "transactions": [CreditCardTransaction(**parse_from_mongo(t)) for t in transactions],
            "total_count": total_count,
            "page": page,
            "page_size": page_size,
            "total_pages": (total_count + page_size - 1) // page_size
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] DAO Exception: {e}")
        print(f"[ERROR] Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/credit-cards/{card_id}/dao")
async def process_card_payment(card_id: str, payment_data: CreditCardTransactionCreate):
    """Process credit card payment (Đáo thẻ) - POS or BILL method"""
    try:
        print(f"[DEBUG] DAO Request - card_id: {card_id}")
        print(f"[DEBUG] payment_data: {payment_data}")
        print(f"[DEBUG] payment_data.dict(): {payment_data.dict()}")
        
        # Validate card exists
        card = await db.credit_cards.find_one({"id": card_id})
        if not card:
            raise HTTPException(status_code=404, detail="Không tìm thấy thẻ")
        
        print(f"[DEBUG] Found card: {card['id']} - {card.get('card_number', 'N/A')}")
        
        # Validate customer exists  
        customer = await db.customers.find_one({"id": card["customer_id"]})
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        print(f"[DEBUG] Found customer: {customer['id']} - {customer.get('name', 'N/A')}")
        
        # Generate transaction group ID
        transaction_group_id = f"CC_{int(datetime.now().timestamp())}"
        print(f"[DEBUG] Generated transaction_group_id: {transaction_group_id}")
        
        if payment_data.payment_method == CreditCardPaymentMethod.POS:
            # POS method - single transaction
            if not payment_data.total_amount or payment_data.total_amount <= 0:
                raise HTTPException(status_code=400, detail="Số tiền đáo không hợp lệ")
            
            total_amount = payment_data.total_amount
            profit_value = round(total_amount * payment_data.profit_pct / 100, 0)
            payback = total_amount - profit_value
            
            # Create single transaction
            transaction = CreditCardTransaction(
                card_id=card_id,
                customer_id=card["customer_id"],
                transaction_group_id=transaction_group_id,
                payment_method=CreditCardPaymentMethod.POS,
                total_amount=total_amount,
                profit_pct=payment_data.profit_pct,
                profit_value=profit_value,
                payback=payback,
                notes=payment_data.notes or f"Đáo thẻ POS - ****{card['card_number'][-4:]}"
            )
            
            # Save transaction
            transaction_dict = prepare_for_mongo(transaction.dict())
            await db.credit_card_transactions.insert_one(transaction_dict)
            
        elif payment_data.payment_method == CreditCardPaymentMethod.BILL:
            # BILL method - multiple transactions from bills
            if not payment_data.bill_ids:
                raise HTTPException(status_code=400, detail="Cần chọn ít nhất một bill điện")
            
            # Get bills and validate they're available
            bill_filter = {"id": {"$in": payment_data.bill_ids}, "status": BillStatus.AVAILABLE}
            bills = await db.bills.find(bill_filter).to_list(None)
            
            if len(bills) != len(payment_data.bill_ids):
                raise HTTPException(status_code=400, detail="Một số bill không khả dụng hoặc không tồn tại")
            
            # Calculate totals
            total_amount = sum(bill.get("amount", 0) for bill in bills)
            profit_value = round(total_amount * payment_data.profit_pct / 100, 0)
            payback = total_amount - profit_value
            
            # Create transactions for each bill (with -1, -2, etc.)
            transactions_created = []
            for i, bill in enumerate(bills):
                bill_transaction_id = f"{transaction_group_id}-{i+1}"
                
                transaction = CreditCardTransaction(
                    id=bill_transaction_id,
                    card_id=card_id,
                    customer_id=card["customer_id"],
                    transaction_group_id=transaction_group_id,
                    payment_method=CreditCardPaymentMethod.BILL,
                    total_amount=bill.get("amount", 0),
                    profit_pct=payment_data.profit_pct,
                    profit_value=round(bill.get("amount", 0) * payment_data.profit_pct / 100, 0),
                    payback=bill.get("amount", 0) - round(bill.get("amount", 0) * payment_data.profit_pct / 100, 0),
                    bill_ids=[bill["id"]],
                    notes=payment_data.notes or f"Đáo thẻ BILL - ****{card['card_number'][-4:]} - {bill.get('customer_code', '')}"
                )
                
                # Save transaction
                transaction_dict = prepare_for_mongo(transaction.dict())
                await db.credit_card_transactions.insert_one(transaction_dict)
                transactions_created.append(transaction)
                
                # Update bill status to SOLD
                await db.bills.update_one(
                    {"id": bill["id"]},
                    {"$set": {"status": BillStatus.SOLD, "updated_at": datetime.now(timezone.utc)}}
                )
            
            # Remove bills from inventory
            await db.inventory_items.delete_many({"bill_id": {"$in": payment_data.bill_ids}})
        
        # Update card cycle tracking and status
        current_date = datetime.now(timezone.utc)
        current_cycle_month = get_current_cycle_month(current_date)
        
        # Get current card data for cycle tracking
        current_card = await db.credit_cards.find_one({"id": card_id})
        cycle_payment_count = current_card.get("cycle_payment_count", 0) + 1
        
        await db.credit_cards.update_one(
            {"id": card_id},
            {"$set": {
                "status": CardStatus.PAID_OFF.value,
                "current_cycle_month": current_cycle_month,
                "last_payment_date": current_date,
                "cycle_payment_count": cycle_payment_count,  # Track multiple payments
                "updated_at": current_date
            }}
        )
        
        # Update customer transaction count
        await db.customers.update_one(
            {"id": card["customer_id"]},
            {"$inc": {"total_transactions": 1}}
        )
        
        # Create corresponding Sale record for customer transaction history
        sale = Sale(
            customer_id=card["customer_id"],
            transaction_type="CREDIT_CARD",
            total=total_amount,
            profit_pct=payment_data.profit_pct,
            profit_value=profit_value,
            payback=payback,
            method=PaymentMethod.OTHER,
            status="COMPLETED",
            notes=f"Đáo thẻ ****{card['card_number'][-4:]} - {payment_data.payment_method.value}",
            bill_ids=payment_data.bill_ids or []
        )
        
        # Save sale record
        sale_dict = prepare_for_mongo(sale.dict())
        await db.sales.insert_one(sale_dict)
        
        # Log activity for dashboard
        activity_type = ActivityType.CARD_PAYMENT_POS if payment_data.payment_method == CreditCardPaymentMethod.POS else ActivityType.CARD_PAYMENT_BILL
        activity_title = f"Đáo thẻ ****{card['card_number'][-4:]} - {format_currency_short(total_amount)} VND"
        
        await log_activity(ActivityCreate(
            type=activity_type,
            title=activity_title,
            description=f"Phương thức: {payment_data.payment_method.value}, Lợi nhuận: {payment_data.profit_pct}%",
            customer_id=card["customer_id"],
            customer_name=customer["name"],
            amount=total_amount,
            status="SUCCESS",
            metadata={
                "card_id": card_id,
                "transaction_group_id": transaction_group_id,
                "method": payment_data.payment_method.value,
                "profit_pct": payment_data.profit_pct
            }
        ))
        
        return {
            "success": True,
            "message": f"Đã đáo thẻ thành công bằng phương thức {payment_data.payment_method.value}",
            "transaction_group_id": transaction_group_id,
            "total_amount": total_amount,
            "profit_value": profit_value,
            "payback": payback
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] DAO Exception: {e}")
        print(f"[ERROR] Exception type: {type(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

# Activity Logging Functions
async def log_activity(activity_data: ActivityCreate):
    """Log system activity for dashboard"""
    try:
        activity = Activity(**activity_data.dict())
        # Manual enum conversion for MongoDB
        activity_dict = activity.dict()
        activity_dict["type"] = activity_dict["type"].value  # Convert enum to string
        activity_dict = prepare_for_mongo(activity_dict)
        await db.activities.insert_one(activity_dict)
        return activity
    except Exception as e:
        print(f"[ERROR] Error logging activity: {e}")
        import traceback
        traceback.print_exc()
        return None

def format_currency_short(amount: float) -> str:
    """Format currency for activity display"""
    if amount >= 1_000_000:
        return f"{amount/1_000_000:.1f}M"
    elif amount >= 1_000:
        return f"{amount/1_000:.0f}K"
    else:
        return f"{amount:.0f}"

@api_router.get("/activities/recent")
async def get_recent_activities(days: int = 3, limit: int = 50):
    """Get recent activities for dashboard"""
    try:
        # Calculate date range (3 days ago)
        since_date = datetime.now(timezone.utc) - timedelta(days=days)
        print(f"[DEBUG] Looking for activities since: {since_date}")
        
        # Get activities within date range
        activities = await db.activities.find({
            "created_at": {"$gte": since_date}
        }).sort("created_at", -1).limit(limit).to_list(limit)
        
        # If no activities in date range, get recent ones anyway (for better UX)
        if len(activities) == 0:
            activities = await db.activities.find({}).sort("created_at", -1).limit(10).to_list(10)
        
        print(f"[DEBUG] Found {len(activities)} activities from database")
        
        # Return raw activities without parsing for now to avoid errors
        result = []
        for activity in activities:
            try:
                # Don't use parse_from_mongo to avoid ObjectId issues
                result.append({
                    "id": activity.get("id"),
                    "type": activity.get("type"),
                    "title": activity.get("title"),
                    "description": activity.get("description"),
                    "customer_id": activity.get("customer_id"),
                    "customer_name": activity.get("customer_name"),
                    "amount": activity.get("amount"),
                    "status": activity.get("status"),
                    "metadata": activity.get("metadata"),
                    "created_at": activity.get("created_at")
                })
            except Exception as e:
                print(f"[ERROR] Error parsing activity: {e}")
        
        print(f"[DEBUG] Returning {len(result)} parsed activities")
        return result
        
    except Exception as e:
        print(f"[ERROR] Error in get_recent_activities: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/credit-cards/initialize-cycles")
async def initialize_credit_card_cycles():
    """Initialize cycle data for existing credit cards"""
    try:
        current_date = datetime.now(timezone.utc)
        current_cycle_month = get_current_cycle_month(current_date)
        
        # Find cards without cycle data
        cards_without_cycle = await db.credit_cards.find({
            "$or": [
                {"current_cycle_month": {"$exists": False}},
                {"current_cycle_month": None}
            ]
        }).to_list(None)
        
        updated_count = 0
        for card in cards_without_cycle:
            # Calculate real-time status
            real_time_status = calculate_card_status_realtime(card, current_date)
            
            # Initialize cycle fields
            update_fields = {
                "current_cycle_month": current_cycle_month,
                "status": real_time_status.value,
                "cycle_payment_count": 0,
                "total_cycles": 1,  # Start with cycle 1
                "updated_at": current_date
            }
            
            # Set last_payment_date only if status is PAID_OFF
            if real_time_status == CardStatus.PAID_OFF:
                update_fields["last_payment_date"] = current_date
            
            await db.credit_cards.update_one(
                {"id": card["id"]},
                {"$set": update_fields}
            )
            updated_count += 1
        
        return {
            "success": True,
            "message": f"Đã khởi tạo chu kỳ cho {updated_count} thẻ",
            "updated_count": updated_count,
            "current_cycle_month": current_cycle_month
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Sales/Transaction APIs
@api_router.post("/sales", response_model=Sale)
async def create_sale(sale_data: SaleCreate):
    """Create new sale transaction"""
    try:
        # Validate customer exists
        customer = await db.customers.find_one({"id": sale_data.customer_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Get bills and validate they're available
        bill_filter = {"id": {"$in": sale_data.bill_ids}, "status": BillStatus.AVAILABLE}
        bills = await db.bills.find(bill_filter).to_list(None)
        
        if len(bills) != len(sale_data.bill_ids):
            raise HTTPException(status_code=400, detail="Một số bill không khả dụng hoặc không tồn tại")
        
        # Calculate totals
        total = sum(bill.get("amount", 0) for bill in bills)
        profit_value = round(total * sale_data.profit_pct / 100, 0)  # Round to VND
        payback = total - profit_value
        
        # Create sale record
        sale = Sale(
            customer_id=sale_data.customer_id,
            transaction_type="ELECTRIC_BILL",
            total=total,
            profit_pct=sale_data.profit_pct,
            profit_value=profit_value,
            payback=payback,
            method=sale_data.method,
            status="COMPLETED",
            notes=sale_data.notes,
            bill_ids=sale_data.bill_ids,
            created_at=datetime.now(timezone.utc)
        )
        
        # Save sale to database
        sale_dict = prepare_for_mongo(sale.dict())
        await db.sales.insert_one(sale_dict)
        
        # Update bill statuses to SOLD
        await db.bills.update_many(
            {"id": {"$in": sale_data.bill_ids}},
            {"$set": {"status": BillStatus.SOLD, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Remove bills from inventory
        await db.inventory_items.delete_many({"bill_id": {"$in": sale_data.bill_ids}})
        
        # Update customer stats
        await db.customers.update_one(
            {"id": sale_data.customer_id},
            {
                "$inc": {
                    "total_transactions": 1,
                    "total_value": total,
                    "total_bills": len(sale_data.bill_ids),
                    "total_profit_generated": profit_value
                },
                "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
            }
        )
        
        # Log activity for bill sale
        try:
            # Get bill codes for activity title
            bill_codes = [bill.get("customer_code", "N/A") for bill in bills]
            bill_codes_str = ", ".join(bill_codes[:3])  # Show first 3 bill codes
            if len(bill_codes) > 3:
                bill_codes_str += f" (+{len(bill_codes)-3} khác)"
            
            activity_title = f"Bán Bill {bill_codes_str} - {format_currency_short(total)} VND"
            
            activity_data = ActivityCreate(
                type=ActivityType.BILL_SALE,
                title=activity_title,
                description=f"Bán {len(sale_data.bill_ids)} bill cho khách hàng {customer.get('name', 'N/A')}",
                customer_id=sale_data.customer_id,
                customer_name=customer.get('name', 'N/A'),
                amount=total,
                status="SUCCESS",
                metadata={
                    "sale_id": sale.id,
                    "bill_count": len(sale_data.bill_ids),
                    "profit_pct": sale_data.profit_pct,
                    "profit_value": profit_value,
                    "method": sale_data.method.value,
                    "bill_codes": bill_codes
                }
            )
            
            await log_activity(activity_data)
        except Exception as e:
            # Don't fail the sale if activity logging fails
            print(f"[WARNING] Failed to log bill sale activity: {e}")
        
        return sale
        
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/sales", response_model=List[Sale])
async def get_sales(
    customer_id: Optional[str] = None,
    page: int = 1,
    page_size: int = 20
):
    """Get sales transactions"""
    try:
        query = {}
        if customer_id:
            query["customer_id"] = customer_id
            
        skip = (page - 1) * page_size
        sales = await db.sales.find(query).skip(skip).limit(page_size).sort("created_at", -1).to_list(page_size)
        
        return [Sale(**parse_from_mongo(sale)) for sale in sales]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/sales/export")
async def export_sales_data():
    """Export sales data to Excel"""
    try:
        # Get all sales (simple version)
        sales = await db.sales.find({}).sort("created_at", -1).to_list(None)
        
        # Create Excel file
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Lịch Sử Bán Bill"
        
        # Headers
        headers = [
            "ID Giao dịch", "Tên khách hàng", "SĐT khách hàng",
            "Loại giao dịch", "Tổng tiền (VND)", "Lợi nhuận (%)",
            "Giá trị lợi nhuận (VND)", "Số tiền trả khách (VND)",
            "Phương thức thanh toán", "Trạng thái", "Ngày giao dịch", "Ghi chú"
        ]
        
        # Write headers with styling
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Write sales data
        for row, sale in enumerate(sales, 2):
            # Get customer info for this sale
            customer = await db.customers.find_one({"id": sale.get("customer_id")})
            customer_name = customer.get("name", "N/A") if customer else "N/A"
            customer_phone = customer.get("phone", "N/A") if customer else "N/A"
            
            worksheet.cell(row=row, column=1, value=sale.get("id"))
            worksheet.cell(row=row, column=2, value=customer_name)
            worksheet.cell(row=row, column=3, value=customer_phone)
            worksheet.cell(row=row, column=4, value=sale.get("transaction_type", "ELECTRIC_BILL"))
            worksheet.cell(row=row, column=5, value=sale.get("total", 0))
            worksheet.cell(row=row, column=6, value=sale.get("profit_pct", 0))
            worksheet.cell(row=row, column=7, value=sale.get("profit_value", 0))
            worksheet.cell(row=row, column=8, value=sale.get("payback", 0))
            worksheet.cell(row=row, column=9, value=sale.get("method", ""))
            worksheet.cell(row=row, column=10, value=sale.get("status", ""))
            worksheet.cell(row=row, column=11, value=sale.get("created_at", ""))
            worksheet.cell(row=row, column=12, value=sale.get("notes", ""))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 25)
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        from fastapi.responses import StreamingResponse
        
        return StreamingResponse(
            BytesIO(excel_buffer.read()),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": "attachment; filename=lich_su_ban_bill.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/sales/{sale_id}", response_model=Sale)
async def get_sale(sale_id: str):
    """Get sale by ID"""
    try:
        sale = await db.sales.find_one({"id": sale_id})
        if not sale:
            raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
        
        return Sale(**parse_from_mongo(sale))
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/sales/stats/summary")
async def get_sales_stats():
    """Get sales statistics"""
    try:
        # Total sales count
        total_sales = await db.sales.count_documents({})
        
        # Total revenue and profit
        pipeline = [
            {
                "$group": {
                    "_id": None,
                    "total_revenue": {"$sum": "$total"},
                    "total_profit": {"$sum": "$profit_value"},
                    "total_payback": {"$sum": "$payback"}
                }
            }
        ]
        
        result = await db.sales.aggregate(pipeline).to_list(1)
        stats = result[0] if result else {"total_revenue": 0, "total_profit": 0, "total_payback": 0}
        
        # Recent sales
        recent_sales = await db.sales.find({}).sort("created_at", -1).limit(5).to_list(5)
        
        return {
            "total_sales": total_sales,
            "total_revenue": stats["total_revenue"],
            "total_profit": stats["total_profit"], 
            "total_payback": stats["total_payback"],
            "recent_sales": [Sale(**parse_from_mongo(sale)) for sale in recent_sales]
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/webhook/checkbill")
async def webhook_checkbill(
    payload: WebhookPayload,
    x_api_key: str = Header(None, alias="X-API-KEY")
):
    """Webhook endpoint for receiving bill check results from FPT/Shopee"""
    try:
        # Log the webhook (in production, verify API key and signature)
        webhook_log = {
            "id": str(uuid.uuid4()),
            "gateway": Gateway.FPT,  # Default to FPT
            "request_id": payload.request_id,
            "payload": payload.dict(),
            "status": "RECEIVED",
            "received_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.webhook_logs.insert_one(webhook_log)
        
        # Process bills
        for bill_data in payload.bills:
            customer_code = bill_data.get("customer_id") or bill_data.get("contractNumber")
            electric_provider = bill_data.get("electric_provider", "mien_nam")
            provider_region = map_provider_region(electric_provider)
            
            # Create/update bill record
            bill_record = {
                "id": str(uuid.uuid4()),
                "gateway": Gateway.FPT,
                "customer_code": customer_code,
                "provider_region": provider_region,
                "provider_name": bill_data.get("provider_name", electric_provider),
                "status": BillStatus.AVAILABLE,
                "meta": bill_data,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.bills.update_one(
                {"customer_code": customer_code, "gateway": Gateway.FPT},
                {"$set": bill_record},
                upsert=True
            )
        
        # Update webhook log status
        await db.webhook_logs.update_one(
            {"request_id": payload.request_id},
            {"$set": {"status": "PROCESSED", "processed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"accepted": True}
    except Exception as e:
        logger.error(f"Webhook processing error: {str(e)}")
        return {"accepted": False, "error": str(e)}

# =============================================================================
# UNIFIED TRANSACTIONS API - TRANG GIAO DỊCH TỔNG HỢP
# =============================================================================

class TransactionType(str, Enum):
    BILL_SALE = "BILL_SALE"           # Bán Bill
    CREDIT_DAO_POS = "CREDIT_DAO_POS" # Đáo Thẻ POS
    CREDIT_DAO_BILL = "CREDIT_DAO_BILL" # Đáo Thẻ BILL

class UnifiedTransaction(BaseModel):
    """Unified transaction model combining Bill Sales and Credit Card DAO"""
    id: str
    type: TransactionType
    customer_id: str
    customer_name: str
    customer_phone: Optional[str] = None
    
    # Transaction details
    total_amount: float
    profit_amount: float
    profit_percentage: float
    payback: Optional[float] = None
    
    # Items involved
    items: List[Dict[str, Any]]  # Bills or Cards
    item_codes: List[str]  # Bill codes or masked card numbers
    item_display: str  # "PB090..., PB091..." or "****1234"
    
    # Payment & Status
    payment_method: Optional[str] = None
    status: str = "COMPLETED"
    notes: Optional[str] = None
    
    # Metadata
    created_at: datetime
    created_by: Optional[str] = None

@api_router.get("/transactions/unified")
async def get_unified_transactions(
    limit: int = 50,
    offset: int = 0,
    transaction_type: Optional[str] = None,
    customer_id: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    search: Optional[str] = None
):
    """Get unified transactions from both Sales and Credit Card DAO"""
    try:
        unified_transactions = []
        
        # Query filters
        match_filters = {}
        if date_from or date_to:
            date_filter = {}
            if date_from:
                date_filter["$gte"] = datetime.fromisoformat(date_from.replace('Z', '+00:00'))
            if date_to:  
                date_filter["$lte"] = datetime.fromisoformat(date_to.replace('Z', '+00:00'))
            match_filters["created_at"] = date_filter
        
        if customer_id:
            match_filters["customer_id"] = customer_id
        
        # Get Bill Sales
        if not transaction_type or transaction_type == "BILL_SALE":
            sales_pipeline = [
                {"$match": match_filters},
                {
                    "$lookup": {
                        "from": "customers",
                        "localField": "customer_id", 
                        "foreignField": "id",
                        "as": "customer"
                    }
                },
                {
                    "$lookup": {
                        "from": "bills",
                        "localField": "bill_ids",
                        "foreignField": "id", 
                        "as": "bills"
                    }
                },
                {"$sort": {"created_at": -1}},
                {"$limit": limit + offset},
                {"$skip": offset}
            ]
            
            sales_cursor = db.sales.aggregate(sales_pipeline)
            sales_results = await sales_cursor.to_list(length=None)
            
            for sale in sales_results:
                customer = sale.get("customer", [{}])[0]
                bills = sale.get("bills", [])
                
                # Create bill codes display
                bill_codes = [bill.get("customer_code", "N/A") for bill in bills]
                item_display = ", ".join(bill_codes[:3])
                if len(bill_codes) > 3:
                    item_display += f" (+{len(bill_codes)-3} khác)"
                
                transaction = UnifiedTransaction(
                    id=sale["id"],
                    type=TransactionType.BILL_SALE,
                    customer_id=sale["customer_id"],
                    customer_name=customer.get("name", "N/A"),
                    customer_phone=customer.get("phone"),
                    total_amount=sale["total"],
                    profit_amount=sale["profit_value"],
                    profit_percentage=sale["profit_pct"],
                    payback=sale.get("payback"),
                    items=[{
                        "id": bill["id"],
                        "code": bill.get("customer_code"),
                        "amount": bill.get("amount"),
                        "type": "BILL"
                    } for bill in bills],
                    item_codes=bill_codes,
                    item_display=item_display,
                    payment_method=sale.get("method"),
                    status=sale.get("status", "COMPLETED"),
                    notes=sale.get("notes"),
                    created_at=sale["created_at"]
                )
                unified_transactions.append(transaction)
        
        # Get Credit Card DAO Transactions
        if not transaction_type or transaction_type in ["CREDIT_DAO_POS", "CREDIT_DAO_BILL"]:
            dao_match_filters = match_filters.copy()
            if transaction_type == "CREDIT_DAO_POS":
                dao_match_filters["payment_method"] = "POS"
            elif transaction_type == "CREDIT_DAO_BILL":
                dao_match_filters["payment_method"] = "BILL"
            
            dao_pipeline = [
                {"$match": dao_match_filters},
                {
                    "$lookup": {
                        "from": "customers",
                        "localField": "customer_id",
                        "foreignField": "id", 
                        "as": "customer"
                    }
                },
                {
                    "$lookup": {
                        "from": "credit_cards",
                        "localField": "card_id",
                        "foreignField": "id",
                        "as": "card"
                    }
                },
                {"$sort": {"created_at": -1}},
                {"$limit": limit + offset},
                {"$skip": offset}
            ]
            
            dao_cursor = db.credit_card_transactions.aggregate(dao_pipeline)
            dao_results = await dao_cursor.to_list(length=None)
            
            for dao in dao_results:
                customer = dao.get("customer", [{}])[0]
                card = dao.get("card", [{}])[0]
                
                # Create masked card number
                card_number = card.get("card_number", "0000")
                masked_card = f"****{card_number[-4:]}"
                
                # Determine transaction type
                tx_type = TransactionType.CREDIT_DAO_POS if dao["payment_method"] == "POS" else TransactionType.CREDIT_DAO_BILL
                
                transaction = UnifiedTransaction(
                    id=dao["id"],
                    type=tx_type,
                    customer_id=dao["customer_id"],
                    customer_name=customer.get("name", "N/A"),
                    customer_phone=customer.get("phone"),
                    total_amount=dao["total_amount"],
                    profit_amount=dao["profit_value"],
                    profit_percentage=dao["profit_pct"],
                    payback=dao.get("payback"),
                    items=[{
                        "id": card["id"],
                        "code": masked_card,
                        "amount": dao["total_amount"],
                        "type": "CREDIT_CARD"
                    }],
                    item_codes=[masked_card],
                    item_display=masked_card,
                    payment_method=dao["payment_method"],
                    status=dao.get("status", "COMPLETED"),
                    notes=dao.get("notes"),
                    created_at=dao["created_at"]
                )
                unified_transactions.append(transaction)
        
        # Sort by created_at descending - handle mixed timezone datetime objects
        def safe_sort_key(transaction):
            created_at = transaction.created_at
            # Convert all datetime objects to timezone-aware UTC for consistent comparison
            if isinstance(created_at, datetime):
                if created_at.tzinfo is None:
                    # Timezone-naive datetime - assume UTC
                    created_at = created_at.replace(tzinfo=timezone.utc)
                else:
                    # Already timezone-aware - convert to UTC
                    created_at = created_at.astimezone(timezone.utc)
            elif isinstance(created_at, str):
                # String datetime - parse and make timezone-aware
                try:
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    if created_at.tzinfo is None:
                        created_at = created_at.replace(tzinfo=timezone.utc)
                except:
                    # Fallback to current time if parsing fails
                    created_at = datetime.now(timezone.utc)
            else:
                # Fallback to current time for unknown types
                created_at = datetime.now(timezone.utc)
            
            return created_at
        
        unified_transactions.sort(key=safe_sort_key, reverse=True)
        
        # Apply search filter if provided
        if search:
            search_lower = search.lower()
            filtered_transactions = []
            for tx in unified_transactions:
                if (search_lower in tx.customer_name.lower() or 
                    search_lower in tx.item_display.lower() or
                    any(search_lower in code.lower() for code in tx.item_codes)):
                    filtered_transactions.append(tx)
            unified_transactions = filtered_transactions
        
        return unified_transactions[:limit]
        
    except Exception as e:
        logger.error(f"Error getting unified transactions: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Transaction Update Models
class SaleUpdate(BaseModel):
    """Model for updating sale transactions"""
    total: Optional[float] = None
    profit_value: Optional[float] = None
    profit_percentage: Optional[float] = None
    notes: Optional[str] = None
    created_at: Optional[datetime] = None

class CreditCardTransactionUpdate(BaseModel):
    """Model for updating credit card transactions"""
    total_amount: Optional[float] = None
    profit_amount: Optional[float] = None
    profit_pct: Optional[float] = None
    notes: Optional[str] = None
    created_at: Optional[datetime] = None

@api_router.put("/transactions/sale/{transaction_id}")
async def update_sale_transaction(transaction_id: str, update_data: SaleUpdate):
    """Update a sale transaction"""
    try:
        # Check if transaction exists
        sale = await db.sales.find_one({"id": transaction_id})
        if not sale:
            raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
        
        # Prepare update data
        update_fields = {}
        if update_data.total is not None:
            update_fields["total"] = update_data.total
        if update_data.profit_value is not None:
            update_fields["profit_value"] = update_data.profit_value
        if update_data.profit_percentage is not None:
            update_fields["profit_percentage"] = update_data.profit_percentage
        if update_data.notes is not None:
            update_fields["notes"] = update_data.notes
        if update_data.created_at is not None:
            update_fields["created_at"] = update_data.created_at
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="Không có dữ liệu để cập nhật")
        
        update_fields["updated_at"] = datetime.now(timezone.utc)
        
        # Update transaction
        result = await db.sales.update_one(
            {"id": transaction_id},
            {"$set": update_fields}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
        
        # Get updated transaction
        updated_sale = await db.sales.find_one({"id": transaction_id})
        # Convert MongoDB document to JSON-serializable format
        updated_sale = parse_from_mongo(updated_sale)
        return {"success": True, "message": "Cập nhật giao dịch thành công", "data": updated_sale}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating sale transaction: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/transactions/credit-card/{transaction_id}")
async def update_credit_card_transaction(transaction_id: str, update_data: CreditCardTransactionUpdate):
    """Update a credit card transaction"""
    try:
        # Check if transaction exists
        credit_tx = await db.credit_card_transactions.find_one({"id": transaction_id})
        if not credit_tx:
            raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch thẻ tín dụng")
        
        # Prepare update data
        update_fields = {}
        if update_data.total_amount is not None:
            update_fields["total_amount"] = update_data.total_amount
        if update_data.profit_amount is not None:
            update_fields["profit_amount"] = update_data.profit_amount
        if update_data.profit_pct is not None:
            update_fields["profit_pct"] = update_data.profit_pct
        if update_data.notes is not None:
            update_fields["notes"] = update_data.notes
        if update_data.created_at is not None:
            update_fields["created_at"] = update_data.created_at
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="Không có dữ liệu để cập nhật")
        
        update_fields["updated_at"] = datetime.now(timezone.utc)
        
        # Update transaction
        result = await db.credit_card_transactions.update_one(
            {"id": transaction_id},
            {"$set": update_fields}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Không tìm thấy giao dịch")
        
        # Get updated transaction
        updated_credit_tx = await db.credit_card_transactions.find_one({"id": transaction_id})
        # Convert MongoDB document to JSON-serializable format
        updated_credit_tx = parse_from_mongo(updated_credit_tx)
        return {"success": True, "message": "Cập nhật giao dịch thẻ tín dụng thành công", "data": updated_credit_tx}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating credit card transaction: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/transactions/stats")
async def get_transactions_stats():
    """Get comprehensive transaction statistics"""
    try:
        # Get current date for filtering
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        this_month = today.replace(day=1)
        
        # Bill Sales Stats
        bill_sales_today = await db.sales.count_documents({
            "created_at": {"$gte": today}
        })
        
        bill_sales_total = await db.sales.count_documents({})
        
        bill_revenue_pipeline = [
            {
                "$group": {
                    "_id": None,
                    "total_revenue": {"$sum": "$total"},
                    "total_profit": {"$sum": "$profit_value"}
                }
            }
        ]
        bill_revenue = await db.sales.aggregate(bill_revenue_pipeline).to_list(1)
        bill_stats = bill_revenue[0] if bill_revenue else {"total_revenue": 0, "total_profit": 0}
        
        # Credit Card DAO Stats
        dao_today = await db.credit_card_transactions.count_documents({
            "created_at": {"$gte": today}
        })
        
        dao_total = await db.credit_card_transactions.count_documents({})
        
        dao_revenue_pipeline = [
            {
                "$group": {
                    "_id": None,
                    "total_revenue": {"$sum": "$total_amount"},
                    "total_profit": {"$sum": "$profit_value"}
                }
            }
        ]
        dao_revenue = await db.credit_card_transactions.aggregate(dao_revenue_pipeline).to_list(1)
        dao_stats = dao_revenue[0] if dao_revenue else {"total_revenue": 0, "total_profit": 0}
        
        # Pending transactions (for future use)
        pending_count = 0  # No pending transactions in current system
        
        return {
            "total_revenue": bill_stats["total_revenue"] + dao_stats["total_revenue"],
            "total_profit": bill_stats["total_profit"] + dao_stats["total_profit"],
            "transactions_today": bill_sales_today + dao_today,
            "bill_sales_count": bill_sales_total,
            "dao_transactions_count": dao_total,
            "pending_count": pending_count,
            "bill_sales_revenue": bill_stats["total_revenue"],
            "bill_sales_profit": bill_stats["total_profit"],
            "dao_revenue": dao_stats["total_revenue"], 
            "dao_profit": dao_stats["total_profit"]
        }
        
    except Exception as e:
        logger.error(f"Error getting transaction stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# REPORTS & ANALYTICS API - TRANG BÁO CÁO
# =============================================================================

@api_router.get("/reports/dashboard-stats")
async def get_dashboard_stats(period: str = "all"):
    """Get comprehensive dashboard statistics for reports - REAL DATA ONLY"""
    try:
        # For now, get all data (will add date filtering later)
        
        # Bill Sales Stats - ALL REAL DATA
        bill_pipeline = [
            {
                "$group": {
                    "_id": None,
                    "total_revenue": {"$sum": "$total"},
                    "total_profit": {"$sum": "$profit_value"},
                    "transaction_count": {"$sum": 1},
                    "avg_transaction": {"$avg": "$total"}
                }
            }
        ]
        
        bill_result = await db.sales.aggregate(bill_pipeline).to_list(1)
        bill_stats = bill_result[0] if bill_result else {
            "total_revenue": 0, "total_profit": 0, "transaction_count": 0, "avg_transaction": 0
        }

        # Credit DAO Stats - ALL REAL DATA
        dao_pipeline = [
            {
                "$group": {
                    "_id": None,
                    "total_revenue": {"$sum": "$total_amount"},
                    "total_profit": {"$sum": "$profit_value"},
                    "transaction_count": {"$sum": 1},
                    "avg_transaction": {"$avg": "$total_amount"}
                }
            }
        ]
        
        dao_result = await db.credit_card_transactions.aggregate(dao_pipeline).to_list(1)
        dao_stats = dao_result[0] if dao_result else {
            "total_revenue": 0, "total_profit": 0, "transaction_count": 0, "avg_transaction": 0
        }

        # Customer Stats - REAL DATA
        total_customers = await db.customers.count_documents({})
        active_customers = await db.customers.count_documents({"is_active": True})

        # Calculate totals
        total_revenue = bill_stats["total_revenue"] + dao_stats["total_revenue"]
        total_profit = bill_stats["total_profit"] + dao_stats["total_profit"]
        total_transactions = bill_stats["transaction_count"] + dao_stats["transaction_count"]
        avg_transaction_value = (bill_stats["avg_transaction"] + dao_stats["avg_transaction"]) / 2 if total_transactions > 0 else 0

        # Calculate profit margin
        profit_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0

        return {
            "period": period,
            "total_revenue": total_revenue,
            "total_profit": total_profit,
            "total_transactions": total_transactions,
            "profit_margin": round(profit_margin, 1),
            "avg_transaction_value": round(avg_transaction_value, 0),
            "customer_stats": {
                "total_customers": total_customers,
                "active_customers": active_customers,
                "inactive_customers": total_customers - active_customers
            },
            "breakdown": {
                "bill_sales": {
                    "revenue": bill_stats["total_revenue"],
                    "profit": bill_stats["total_profit"],
                    "transactions": bill_stats["transaction_count"],
                    "avg_value": round(bill_stats["avg_transaction"], 0)
                },
                "credit_dao": {
                    "revenue": dao_stats["total_revenue"],
                    "profit": dao_stats["total_profit"],
                    "transactions": dao_stats["transaction_count"],
                    "avg_value": round(dao_stats["avg_transaction"], 0)
                }
            }
        }

    except Exception as e:
        logger.error(f"Error getting dashboard stats: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/revenue-trend")
async def get_revenue_trend(months: int = 6):
    """Get revenue trend data for charts - real data only"""
    try:
        # Calculate date range for the specified months
        end_date = datetime.now(timezone.utc)
        start_date = end_date - timedelta(days=months * 30)  # Approximate months to days
        
        # Create monthly buckets
        monthly_data = []
        current_date = start_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        while current_date < end_date:
            # Calculate next month
            if current_date.month == 12:
                next_month = current_date.replace(year=current_date.year + 1, month=1)
            else:
                next_month = current_date.replace(month=current_date.month + 1)
            
            month_filter = {
                "created_at": {
                    "$gte": current_date,
                    "$lt": next_month
                }
            }
            
            # Bill sales for this month
            bill_pipeline = [
                {"$match": month_filter},
                {
                    "$group": {
                        "_id": None,
                        "revenue": {"$sum": "$total"},
                        "profit": {"$sum": "$profit_value"},
                        "count": {"$sum": 1}
                    }
                }
            ]
            
            bill_result = await db.sales.aggregate(bill_pipeline).to_list(1)
            bill_data = bill_result[0] if bill_result else {"revenue": 0, "profit": 0, "count": 0}
            
            # Credit DAO for this month
            dao_pipeline = [
                {"$match": month_filter},
                {
                    "$group": {
                        "_id": None,
                        "revenue": {"$sum": "$total_amount"},
                        "profit": {"$sum": "$profit_value"},
                        "count": {"$sum": 1}
                    }
                }
            ]
            
            dao_result = await db.credit_card_transactions.aggregate(dao_pipeline).to_list(1)
            dao_data = dao_result[0] if dao_result else {"revenue": 0, "profit": 0, "count": 0}
            
            monthly_data.append({
                "month": current_date.strftime("%Y-%m"),
                "month_name": current_date.strftime("%m/%Y"),
                "total_revenue": bill_data["revenue"] + dao_data["revenue"],
                "total_profit": bill_data["profit"] + dao_data["profit"],
                "total_transactions": bill_data["count"] + dao_data["count"],
                "bill_sales_revenue": bill_data["revenue"],
                "dao_revenue": dao_data["revenue"],
                "bill_sales_count": bill_data["count"],
                "dao_count": dao_data["count"]
            })
            
            current_date = next_month
        
        return {
            "period": f"{months} months",
            "data": monthly_data
        }

    except Exception as e:
        logger.error(f"Error getting revenue trend: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/customer-analytics")
async def get_customer_analytics():
    """Get customer analytics - real data from database"""
    try:
        # Customer distribution by type
        customer_type_pipeline = [
            {
                "$group": {
                    "_id": "$type",
                    "count": {"$sum": 1}
                }
            }
        ]
        
        customer_types = await db.customers.aggregate(customer_type_pipeline).to_list(10)
        
        # Top customers by transaction value
        top_customers_pipeline = [
            {
                "$lookup": {
                    "from": "sales",
                    "localField": "id",
                    "foreignField": "customer_id",
                    "as": "sales"
                }
            },
            {
                "$lookup": {
                    "from": "credit_card_transactions",
                    "localField": "id",
                    "foreignField": "customer_id", 
                    "as": "dao_transactions"
                }
            },
            {
                "$addFields": {
                    "total_sales_value": {"$sum": "$sales.total"},
                    "total_dao_value": {"$sum": "$dao_transactions.total_amount"},
                    "sales_count": {"$size": "$sales"},
                    "dao_count": {"$size": "$dao_transactions"}
                }
            },
            {
                "$addFields": {
                    "total_value": {"$add": ["$total_sales_value", "$total_dao_value"]},
                    "total_transactions": {"$add": ["$sales_count", "$dao_count"]}
                }
            },
            {"$sort": {"total_value": -1}},
            {"$limit": 10},
            {
                "$project": {
                    "name": 1,
                    "phone": 1,
                    "type": 1,
                    "total_value": 1,
                    "total_transactions": 1,
                    "total_sales_value": 1,
                    "total_dao_value": 1
                }
            }
        ]
        
        top_customers = await db.customers.aggregate(top_customers_pipeline).to_list(10)
        
        # Customer activity distribution
        customer_activity_pipeline = [
            {
                "$group": {
                    "_id": "$is_active",
                    "count": {"$sum": 1}
                }
            }
        ]
        
        customer_activity = await db.customers.aggregate(customer_activity_pipeline).to_list(10)
        
        return {
            "customer_distribution": customer_types,
            "top_customers": top_customers,
            "activity_distribution": customer_activity,
            "total_customers": await db.customers.count_documents({})
        }

    except Exception as e:
        logger.error(f"Error getting customer analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/charts/revenue-trend")
async def get_revenue_trend_chart(months: int = 12):
    """Get revenue trend data for line chart - REAL DATA ONLY"""
    try:
        # Get all sales transactions
        sales_cursor = db.sales.find({}).sort("created_at", 1)
        sales_data = await sales_cursor.to_list(None)
        
        # Get all DAO transactions  
        dao_cursor = db.credit_card_transactions.find({}).sort("created_at", 1)
        dao_data = await dao_cursor.to_list(None)
        
        # Group data by month
        months_data = {}
        
        # Process sales data
        for sale in sales_data:
            if sale.get("created_at"):
                try:
                    # Parse datetime string
                    dt = datetime.fromisoformat(sale["created_at"].replace('Z', '+00:00'))
                    month_key = dt.strftime("%Y-%m")
                except:
                    continue
                    
                if month_key not in months_data:
                    months_data[month_key] = {
                        "month": month_key,
                        "sales_revenue": 0,
                        "sales_profit": 0,
                        "sales_count": 0,
                        "dao_revenue": 0,
                        "dao_profit": 0,
                        "dao_count": 0
                    }
                months_data[month_key]["sales_revenue"] += sale.get("total", 0)
                months_data[month_key]["sales_profit"] += sale.get("profit_value", 0)
                months_data[month_key]["sales_count"] += 1
        
        # Process DAO data
        for dao in dao_data:
            if dao.get("created_at"):
                try:
                    # Parse datetime string
                    dt = datetime.fromisoformat(dao["created_at"].replace('Z', '+00:00'))
                    month_key = dt.strftime("%Y-%m")
                except:
                    continue
                    
                if month_key not in months_data:
                    months_data[month_key] = {
                        "month": month_key,
                        "sales_revenue": 0,
                        "sales_profit": 0,
                        "sales_count": 0,
                        "dao_revenue": 0,
                        "dao_profit": 0,
                        "dao_count": 0
                    }
                months_data[month_key]["dao_revenue"] += dao.get("total_amount", 0)
                months_data[month_key]["dao_profit"] += dao.get("profit_value", 0)
                months_data[month_key]["dao_count"] += 1
        
        # Format for chart (last N months)
        chart_data = []
        for month_key in sorted(months_data.keys())[-months:]:
            data = months_data[month_key]
            total_revenue = data["sales_revenue"] + data["dao_revenue"]
            total_profit = data["sales_profit"] + data["dao_profit"]
            total_transactions = data["sales_count"] + data["dao_count"]
            
            # Convert month to display format
            try:
                month_display = datetime.strptime(month_key, "%Y-%m").strftime("%m/%Y")
            except:
                month_display = month_key
            
            chart_data.append({
                "month": month_key,
                "month_display": month_display,
                "total_revenue": total_revenue,
                "total_profit": total_profit,
                "total_transactions": total_transactions,
                "sales_revenue": data["sales_revenue"],
                "dao_revenue": data["dao_revenue"],
                "profit_margin": round((total_profit / total_revenue * 100) if total_revenue > 0 else 0, 1)
            })
        
        return {
            "success": True,
            "data": chart_data,
            "summary": {
                "total_months": len(chart_data),
                "total_revenue": sum(item["total_revenue"] for item in chart_data),
                "total_profit": sum(item["total_profit"] for item in chart_data),
                "avg_monthly_revenue": sum(item["total_revenue"] for item in chart_data) / len(chart_data) if chart_data else 0
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting revenue trend chart: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/charts/transaction-distribution")
async def get_transaction_distribution_chart():
    """Get transaction distribution for pie chart - REAL DATA ONLY"""
    try:
        # Get bill sales stats
        bill_stats = await db.sales.aggregate([
            {
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1},
                    "revenue": {"$sum": "$total"},
                    "profit": {"$sum": "$profit_value"}
                }
            }
        ]).to_list(1)
        
        bill_data = bill_stats[0] if bill_stats else {"count": 0, "revenue": 0, "profit": 0}
        
        # Get DAO stats by payment method
        dao_stats = await db.credit_card_transactions.aggregate([
            {
                "$group": {
                    "_id": "$payment_method",
                    "count": {"$sum": 1},
                    "revenue": {"$sum": "$total_amount"},
                    "profit": {"$sum": "$profit_value"}
                }
            }
        ]).to_list(10)
        
        # Process DAO data
        dao_pos = {"count": 0, "revenue": 0, "profit": 0}
        dao_bill = {"count": 0, "revenue": 0, "profit": 0}
        
        for item in dao_stats:
            if item["_id"] == "POS":
                dao_pos = {"count": item["count"], "revenue": item["revenue"], "profit": item["profit"]}
            elif item["_id"] == "BILL":
                dao_bill = {"count": item["count"], "revenue": item["revenue"], "profit": item["profit"]}
        
        # Calculate total for percentages
        total_revenue = bill_data["revenue"] + dao_pos["revenue"] + dao_bill["revenue"]
        total_transactions = bill_data["count"] + dao_pos["count"] + dao_bill["count"]
        
        # Format for pie chart
        chart_data = [
            {
                "name": "Bán Bill",
                "value": bill_data["revenue"],
                "count": bill_data["count"],
                "profit": bill_data["profit"],
                "percentage": round((bill_data["revenue"] / total_revenue * 100) if total_revenue > 0 else 0, 1),
                "color": "#10B981"  # Green
            },
            {
                "name": "Đáo Thẻ POS", 
                "value": dao_pos["revenue"],
                "count": dao_pos["count"],
                "profit": dao_pos["profit"],
                "percentage": round((dao_pos["revenue"] / total_revenue * 100) if total_revenue > 0 else 0, 1),
                "color": "#3B82F6"  # Blue
            },
            {
                "name": "Đáo Thẻ BILL",
                "value": dao_bill["revenue"],
                "count": dao_bill["count"], 
                "profit": dao_bill["profit"],
                "percentage": round((dao_bill["revenue"] / total_revenue * 100) if total_revenue > 0 else 0, 1),
                "color": "#8B5CF6"  # Purple
            }
        ]
        
        # Filter out zero values
        chart_data = [item for item in chart_data if item["value"] > 0]
        
        return {
            "success": True,
            "data": chart_data,
            "summary": {
                "total_revenue": total_revenue,
                "total_transactions": total_transactions,
                "largest_segment": max(chart_data, key=lambda x: x["value"])["name"] if chart_data else None
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting transaction distribution chart: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/reports/charts/top-customers")
async def get_top_customers_chart(limit: int = 10):
    """Get top customers for bar chart - REAL DATA ONLY"""
    try:
        # Aggregate customer data from both sales and DAO transactions
        pipeline = [
            {
                "$lookup": {
                    "from": "sales",
                    "localField": "id",
                    "foreignField": "customer_id",
                    "as": "sales"
                }
            },
            {
                "$lookup": {
                    "from": "credit_card_transactions",
                    "localField": "id", 
                    "foreignField": "customer_id",
                    "as": "dao_transactions"
                }
            },
            {
                "$addFields": {
                    "total_sales_revenue": {"$sum": "$sales.total"},
                    "total_dao_revenue": {"$sum": "$dao_transactions.total_amount"},
                    "total_sales_profit": {"$sum": "$sales.profit_value"},
                    "total_dao_profit": {"$sum": "$dao_transactions.profit_value"},
                    "sales_count": {"$size": "$sales"},
                    "dao_count": {"$size": "$dao_transactions"}
                }
            },
            {
                "$addFields": {
                    "total_revenue": {"$add": ["$total_sales_revenue", "$total_dao_revenue"]},
                    "total_profit": {"$add": ["$total_sales_profit", "$total_dao_profit"]},
                    "total_transactions": {"$add": ["$sales_count", "$dao_count"]}
                }
            },
            {
                "$match": {
                    "total_revenue": {"$gt": 0}
                }
            },
            {"$sort": {"total_revenue": -1}},
            {"$limit": limit},
            {
                "$project": {
                    "name": 1,
                    "phone": 1,
                    "type": 1,
                    "total_revenue": 1,
                    "total_profit": 1,
                    "total_transactions": 1,
                    "total_sales_revenue": 1,
                    "total_dao_revenue": 1
                }
            }
        ]
        
        customers = await db.customers.aggregate(pipeline).to_list(limit)
        
        # Format for bar chart
        chart_data = []
        for i, customer in enumerate(customers):
            chart_data.append({
                "rank": i + 1,
                "name": customer["name"],
                "phone": customer.get("phone", "N/A"),
                "type": customer.get("type", "N/A"),
                "total_revenue": customer["total_revenue"],
                "total_profit": customer["total_profit"],
                "total_transactions": customer["total_transactions"],
                "sales_revenue": customer["total_sales_revenue"],
                "dao_revenue": customer["total_dao_revenue"],
                "profit_margin": round((customer["total_profit"] / customer["total_revenue"] * 100) if customer["total_revenue"] > 0 else 0, 1)
            })
        
        return {
            "success": True,
            "data": chart_data,
            "summary": {
                "total_customers": len(chart_data),
                "top_customer_revenue": chart_data[0]["total_revenue"] if chart_data else 0,
                "avg_customer_revenue": sum(item["total_revenue"] for item in chart_data) / len(chart_data) if chart_data else 0
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting top customers chart: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# =============================================================================
# CUSTOMER DETAIL PAGE APIs - 360° Customer View
# =============================================================================

@api_router.get("/customers/{customer_id}/detailed-profile")
async def get_customer_detailed_profile(customer_id: str):
    """Get comprehensive customer profile with all related data - supports both UUID and ObjectId lookup"""
    try:
        # Try to find customer by 'id' field first (UUID format)
        customer = await db.customers.find_one({"id": customer_id})
        
        # If not found and customer_id looks like ObjectId, try _id field
        if not customer and len(customer_id) == 24 and all(c in '0123456789abcdef' for c in customer_id.lower()):
            try:
                from bson import ObjectId
                customer = await db.customers.find_one({"_id": ObjectId(customer_id)})
                # If found by ObjectId, use the actual 'id' field for subsequent queries
                if customer and customer.get('id'):
                    customer_id = customer.get('id')
            except:
                pass  # Invalid ObjectId format, continue with original customer_id
        
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Get customer's credit cards
        cards_cursor = db.credit_cards.find({"customer_id": customer_id})
        cards = await cards_cursor.to_list(None)
        
        # Get customer's bill sales
        sales_pipeline = [
            {"$match": {"customer_id": customer_id}},
            {
                "$lookup": {
                    "from": "bills",
                    "localField": "bill_ids", 
                    "foreignField": "id",
                    "as": "bills"
                }
            },
            {"$sort": {"created_at": -1}}
        ]
        sales_cursor = db.sales.aggregate(sales_pipeline)
        sales = await sales_cursor.to_list(None)
        
        # Get customer's DAO transactions
        dao_cursor = db.credit_card_transactions.find({"customer_id": customer_id}).sort("created_at", -1)
        dao_transactions = await dao_cursor.to_list(None)
        
        # Calculate customer metrics
        total_sales_value = sum(sale.get("total", 0) for sale in sales)
        total_sales_profit = sum(sale.get("profit_value", 0) for sale in sales)
        total_dao_value = sum(dao.get("total_amount", 0) for dao in dao_transactions)
        total_dao_profit = sum(dao.get("profit_value", 0) for dao in dao_transactions)
        
        total_transaction_value = total_sales_value + total_dao_value
        total_profit = total_sales_profit + total_dao_profit
        total_transactions = len(sales) + len(dao_transactions)
        
        avg_transaction_value = total_transaction_value / total_transactions if total_transactions > 0 else 0
        profit_margin = (total_profit / total_transaction_value * 100) if total_transaction_value > 0 else 0
        
        # Customer tier calculation
        if total_transaction_value >= 50000000:  # 50M VND
            tier = "VIP"
        elif total_transaction_value >= 20000000:  # 20M VND  
            tier = "Premium"
        elif total_transaction_value >= 5000000:   # 5M VND
            tier = "Regular"
        else:
            tier = "New"
        
        # Calculate credit cards metrics
        total_credit_limit = sum(card.get("credit_limit", 0) for card in cards)
        active_cards = [card for card in cards if card.get("status") != "Hết hạn"]
        
        # Recent activity (last 10 transactions)
        recent_activities = []
        
        # Add recent sales
        for sale in sales[:5]:
            recent_activities.append({
                "id": sale["id"],
                "type": "BILL_SALE",
                "amount": sale.get("total", 0),
                "profit": sale.get("profit_value", 0),
                "created_at": sale["created_at"],
                "description": f"Bán {len(sale.get('bills', []))} bills",
                "bills_count": len(sale.get("bills", []))
            })
        
        # Add recent DAO transactions
        for dao in dao_transactions[:5]:
            card_number = "N/A"
            # Find card number from cards list
            for card in cards:
                if card["id"] == dao.get("card_id"):
                    card_number = f"****{card.get('card_number', '0000')[-4:]}"
                    break
                    
            dao_type = "CREDIT_DAO_POS" if dao.get("payment_method", "POS") == "POS" else "CREDIT_DAO_BILL"
            recent_activities.append({
                "id": dao["id"],
                "type": dao_type,
                "amount": dao.get("total_amount", 0),
                "profit": dao.get("profit_value", 0),
                "created_at": dao["created_at"],
                "description": f"Đáo Thẻ {dao.get('payment_method', 'POS')} {card_number}",
                "card_number": card_number
            })
        
        # Sort recent activities by date - handle mixed timezone datetime objects
        def safe_activity_sort_key(activity):
            created_at = activity.get("created_at")
            # Convert all datetime objects to timezone-aware UTC for consistent comparison
            if isinstance(created_at, datetime):
                if created_at.tzinfo is None:
                    # Timezone-naive datetime - assume UTC
                    created_at = created_at.replace(tzinfo=timezone.utc)
                else:
                    # Already timezone-aware - convert to UTC
                    created_at = created_at.astimezone(timezone.utc)
            elif isinstance(created_at, str):
                # String datetime - parse and make timezone-aware
                try:
                    created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    if created_at.tzinfo is None:
                        created_at = created_at.replace(tzinfo=timezone.utc)
                except:
                    # Fallback to current time if parsing fails
                    created_at = datetime.now(timezone.utc)
            else:
                # Fallback to current time for unknown types
                created_at = datetime.now(timezone.utc)
            
            return created_at
        
        recent_activities.sort(key=safe_activity_sort_key, reverse=True)
        recent_activities = recent_activities[:10]
        
        return {
            "success": True,
            "customer": {
                "id": customer["id"],
                "name": customer["name"],
                "phone": customer.get("phone"),
                "email": customer.get("email"),
                "address": customer.get("address"),
                "type": customer.get("type", "INDIVIDUAL"),
                "is_active": customer.get("is_active", True),
                "created_at": customer["created_at"],
                "tier": tier,
                "notes": customer.get("notes", "")
            },
            "metrics": {
                "total_transaction_value": total_transaction_value,
                "total_profit": total_profit,
                "total_transactions": total_transactions,
                "avg_transaction_value": avg_transaction_value,
                "profit_margin": round(profit_margin, 1),
                "sales_transactions": len(sales),
                "dao_transactions": len(dao_transactions),
                "sales_value": total_sales_value,
                "dao_value": total_dao_value
            },
            "credit_cards": {
                "total_cards": len(cards),
                "active_cards": len(active_cards),
                "total_credit_limit": total_credit_limit,
                "cards": [{
                    "id": card["id"],
                    "card_number": f"****{card.get('card_number', '0000')[-4:]}",
                    "bank_name": card.get("bank_name"),
                    "card_type": card.get("card_type"),
                    "credit_limit": card.get("credit_limit", 0),
                    "status": card.get("status"),
                    "expiry_date": card.get("expiry_date")
                } for card in cards]
            },
            "recent_activities": recent_activities,
            "performance": {
                "best_month_revenue": 0,  # TODO: Calculate from monthly data
                "growth_rate": 0,         # TODO: Calculate YoY growth
                "success_rate": 100,      # TODO: Calculate based on transaction success
                "avg_days_between_transactions": 0  # TODO: Calculate
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error getting customer detailed profile: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/customers/{customer_id}/transactions-summary")
async def get_customer_transactions_summary(customer_id: str, limit: int = 50):
    """Get customer's transaction history with pagination"""
    try:
        transactions = []
        
        # Get customer's bill sales
        sales_pipeline = [
            {"$match": {"customer_id": customer_id}},
            {
                "$lookup": {
                    "from": "bills",
                    "localField": "bill_ids",
                    "foreignField": "id", 
                    "as": "bills"
                }
            },
            {"$sort": {"created_at": -1}},
            {"$limit": limit}
        ]
        
        sales_cursor = db.sales.aggregate(sales_pipeline)
        sales = await sales_cursor.to_list(None)
        
        for sale in sales:
            bills = sale.get("bills", [])
            bill_codes = [bill.get("customer_code", "N/A") for bill in bills]
            
            transactions.append({
                "id": sale["id"],
                "type": "BILL_SALE",
                "type_display": "Bán Bill",
                "amount": sale.get("total", 0),
                "profit": sale.get("profit_value", 0),
                "created_at": sale["created_at"],
                "description": f"{len(bills)} bills: {', '.join(bill_codes[:3])}{'...' if len(bill_codes) > 3 else ''}",
                "items_count": len(bills),
                "method": sale.get("method"),
                "status": "COMPLETED"
            })
        
        # Get customer's DAO transactions
        dao_cursor = db.credit_card_transactions.find({"customer_id": customer_id}).sort("created_at", -1).limit(limit)
        dao_transactions = await dao_cursor.to_list(None)
        
        # Get cards for card number mapping
        cards_cursor = db.credit_cards.find({"customer_id": customer_id})
        cards = await cards_cursor.to_list(None)
        cards_map = {card["id"]: card for card in cards}
        
        for dao in dao_transactions:
            card_info = cards_map.get(dao.get("card_id"), {})
            card_number = f"****{card_info.get('card_number', '0000')[-4:]}"
            
            dao_type = "CREDIT_DAO_POS" if dao.get("payment_method", "POS") == "POS" else "CREDIT_DAO_BILL"
            transactions.append({
                "id": dao["id"],
                "type": dao_type,
                "type_display": f"Đáo Thẻ {dao.get('payment_method', 'POS')}",
                "amount": dao.get("total_amount", 0),
                "profit": dao.get("profit_value", 0),
                "created_at": dao["created_at"],
                "description": f"{card_info.get('bank_name', 'Unknown')} {card_number}",
                "card_number": card_number,
                "method": dao.get("payment_method"),
                "status": "COMPLETED"
            })
        
        # Sort all transactions by date
        transactions.sort(key=lambda x: x["created_at"], reverse=True)
        
        return {
            "success": True,
            "transactions": transactions[:limit],
            "total_count": len(transactions)
        }
        
    except Exception as e:
        logger.error(f"Error getting customer transactions summary: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/customers/{customer_id}/analytics")
async def get_customer_analytics(customer_id: str):
    """Get customer analytics and performance insights"""
    try:
        # Get customer data
        customer = await db.customers.find_one({"id": customer_id})
        if not customer:
            raise HTTPException(status_code=404, detail="Không tìm thấy khách hàng")
        
        # Monthly transaction analysis
        sales_monthly = await db.sales.aggregate([
            {"$match": {"customer_id": customer_id}},
            {
                "$group": {
                    "_id": {
                        "$dateToString": {
                            "format": "%Y-%m",
                            "date": "$created_at"
                        }
                    },
                    "revenue": {"$sum": "$total"},
                    "profit": {"$sum": "$profit_value"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id": 1}}
        ]).to_list(None)
        
        dao_monthly = await db.credit_card_transactions.aggregate([
            {"$match": {"customer_id": customer_id}},
            {
                "$group": {
                    "_id": {
                        "$dateToString": {
                            "format": "%Y-%m",
                            "date": "$created_at"
                        }
                    },
                    "revenue": {"$sum": "$total_amount"},
                    "profit": {"$sum": "$profit_value"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id": 1}}
        ]).to_list(None)
        
        # Combine monthly data
        monthly_data = {}
        for item in sales_monthly:
            month = item["_id"]
            monthly_data[month] = {
                "month": month,
                "sales_revenue": item["revenue"],
                "sales_profit": item["profit"],
                "sales_count": item["count"],
                "dao_revenue": 0,
                "dao_profit": 0,
                "dao_count": 0
            }
        
        for item in dao_monthly:
            month = item["_id"]
            if month in monthly_data:
                monthly_data[month]["dao_revenue"] = item["revenue"]
                monthly_data[month]["dao_profit"] = item["profit"]
                monthly_data[month]["dao_count"] = item["count"]
            else:
                monthly_data[month] = {
                    "month": month,
                    "sales_revenue": 0,
                    "sales_profit": 0,
                    "sales_count": 0,
                    "dao_revenue": item["revenue"],
                    "dao_profit": item["profit"],
                    "dao_count": item["count"]
                }
        
        # Format monthly chart data
        monthly_chart = []
        for month_key in sorted(monthly_data.keys()):
            data = monthly_data[month_key]
            total_revenue = data["sales_revenue"] + data["dao_revenue"]
            total_profit = data["sales_profit"] + data["dao_profit"]
            
            monthly_chart.append({
                "month": month_key,
                "month_display": datetime.strptime(month_key, "%Y-%m").strftime("%m/%Y"),
                "total_revenue": total_revenue,
                "total_profit": total_profit,
                "sales_revenue": data["sales_revenue"],
                "dao_revenue": data["dao_revenue"],
                "transaction_count": data["sales_count"] + data["dao_count"]
            })
        
        return {
            "success": True,
            "monthly_trend": monthly_chart,
            "insights": {
                "best_month": max(monthly_chart, key=lambda x: x["total_revenue"]) if monthly_chart else None,
                "avg_monthly_revenue": sum(item["total_revenue"] for item in monthly_chart) / len(monthly_chart) if monthly_chart else 0,
                "growth_trend": "stable",  # TODO: Calculate actual trend
                "preferred_transaction_type": "bills" if len(sales_monthly) > len(dao_monthly) else "dao"
            }
        }
        
    except Exception as e:
        logger.error(f"Error getting customer analytics: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# Seed some sample data on startup
@app.on_event("startup")
async def seed_data():
    """Seed sample data for development"""
    try:
        # Check if data already exists
        existing_customers = await db.customers.count_documents({})
        if existing_customers > 0:
            return
            
        # Create sample customers
        sample_customers = [
            {
                "id": str(uuid.uuid4()),
                "type": CustomerType.INDIVIDUAL,
                "name": "Nguyễn Văn A",
                "phone": "0901234567",
                "email": "nguyenvana@example.com",
                "address": "123 Đường ABC, Quận 1, TP.HCM",
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "type": CustomerType.INDIVIDUAL,
                "name": "Trần Thị B",
                "phone": "0907654321",
                "email": "tranthib@example.com",
                "address": "456 Đường XYZ, Quận 2, TP.HCM",
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "type": CustomerType.AGENT,
                "name": "Công ty TNHH ABC",
                "phone": "0281234567",
                "email": "info@abc.com",
                "address": "789 Đường DEF, Quận 3, TP.HCM",
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        
        await db.customers.insert_many(sample_customers)
        
        # Create sample bills for inventory
        sample_bills = [
            {
                "id": str(uuid.uuid4()),
                "gateway": Gateway.FPT,
                "customer_code": "PA22040501111",
                "provider_region": ProviderRegion.MIEN_NAM,
                "provider_name": "Điện lực miền Nam",
                "full_name": "Lê Thị Mai",
                "address": "111 Nguyễn Trãi, Quận 5, TP.HCM",
                "amount": 850000,
                "billing_cycle": "08/2025",
                "status": BillStatus.AVAILABLE,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "gateway": Gateway.FPT,
                "customer_code": "PA22040502222",
                "provider_region": ProviderRegion.MIEN_BAC,
                "provider_name": "Điện lực miền Bắc",
                "full_name": "Hoàng Văn Nam",
                "address": "222 Láng Hạ, Ba Đình, Hà Nội",
                "amount": 1200000,
                "billing_cycle": "08/2025",
                "status": BillStatus.AVAILABLE,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "gateway": Gateway.FPT,
                "customer_code": "PA22040503333",
                "provider_region": ProviderRegion.HCMC,
                "provider_name": "Điện lực TP.HCM",
                "full_name": "Phan Minh Đức",
                "address": "333 Lê Văn Sỹ, Quận 3, TP.HCM",
                "amount": 950000,
                "billing_cycle": "08/2025",
                "status": BillStatus.AVAILABLE,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "gateway": Gateway.FPT,
                "customer_code": "PA22040504444",
                "provider_region": ProviderRegion.MIEN_NAM,
                "provider_name": "Điện lực miền Nam",
                "full_name": "Võ Thị Lan",
                "address": "444 Cách Mạng Tháng 8, Quận 10, TP.HCM",
                "amount": 750000,
                "billing_cycle": "07/2025",
                "status": BillStatus.AVAILABLE,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "gateway": Gateway.FPT,
                "customer_code": "PA22040505555",
                "provider_region": ProviderRegion.MIEN_BAC,
                "provider_name": "Điện lực miền Bắc",
                "full_name": "Đặng Quang Hải",
                "address": "555 Giải Phóng, Hai Bà Trưng, Hà Nội",
                "amount": 1450000,
                "billing_cycle": "08/2025",
                "status": BillStatus.PENDING,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            },
            {
                "id": str(uuid.uuid4()),
                "gateway": Gateway.FPT,
                "customer_code": "PA22040506666",
                "provider_region": ProviderRegion.HCMC,
                "provider_name": "Điện lực TP.HCM",
                "full_name": "Bùi Thị Hồng",
                "address": "666 Phan Xích Long, Phú Nhuận, TP.HCM",
                "amount": 680000,
                "billing_cycle": "07/2025",
                "status": BillStatus.SOLD,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        ]
        
        await db.bills.insert_many(sample_bills)
        
        # Create inventory items for some bills
        inventory_batch_id = f"batch_{str(uuid.uuid4())[:8]}_initial_import"
        inventory_items = []
        
        for i, bill in enumerate(sample_bills[:4]):  # Add first 4 bills to inventory
            inventory_item = {
                "id": str(uuid.uuid4()),
                "bill_id": bill["id"],
                "note": "Import ban đầu từ hệ thống" if i < 2 else "Batch kiểm tra tháng 08",
                "batch_id": inventory_batch_id if i < 2 else f"batch_{str(uuid.uuid4())[:8]}_test_08",
                "added_by": "system",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            inventory_items.append(inventory_item)
        
        if inventory_items:
            await db.inventory_items.insert_many(inventory_items)
        
        logger.info("Sample data seeded successfully - customers, bills, and inventory")
        
    except Exception as e:
        logger.error(f"Error seeding data: {str(e)}")

# Include the router in the main app (must be after all route definitions)
app.include_router(api_router)